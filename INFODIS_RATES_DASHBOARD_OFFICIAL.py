import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font,Color
import win32com.client
import zipfile
import os



# check today date
import datetime
print(datetime.datetime.now())
today = datetime.date.today()

# asking if regular update is needed or it is outside normal working routine
reporting = input('regular update (y) or other date (provide YYYY-MM-DD)? ')

if len(reporting)<5:
    # checking the name of the today
    checking_date = datetime.datetime.today().strftime("%A")

    # if it is not monday it takes one day before
    if  checking_date == 'Monday':
        datum = datetime.datetime.today() - datetime.timedelta(3)
        datum = datum.strftime('%Y-%m-%d')
    # if it is monday it takes three days before
    else:
        datum = datetime.datetime.today() - datetime.timedelta(1)
        datum = datum.strftime('%Y-%m-%d')
else:
    datum = reporting

print(datum)
# naming file
naming_file = r'C:\Users\310295192\Desktop\Work\Rates\Dashboard\reports\rates infodis {}.csv'.format(today)


#getting to inbox
folder = win32com.client.Dispatch("Outlook.Application").GetNameSpace("MAPI").GetDefaultFolder(6)
#getting to infodis folder
subfolder=folder.Folders(3)
#getting to all emails
email = subfolder.Items
#checking the number of emails to further take this value into the loop
x=len(email)
#sort all emails that we have in the folder
email.Sort('ReceivedTime')
first=x-7
#starts the loop through the emails
for infodis in range(8):
#checks the number of email after sorting and extract data on it
    message = email.Item(first + infodis)
    subjectofemail=message.subject
    #now we will be checking each email subject and save it in proper folder
    print(subjectofemail)
    if subjectofemail == "Complex report - missing rates":
        attachment = message.Attachments.Item(1)
        # if the email is for FCL missing rates - it saves it in respective folder
        attachment.SaveAsFile(r'C:\Users\310295192\Desktop\SHIT HAPPENING\Ruurd files\complex_report.zip')

# first extraction
path_to_zip_file = r'C:\Users\310295192\Desktop\SHIT HAPPENING\Ruurd files\complex_report.zip'
directory_to_extract_to = r'C:\Users\310295192\Desktop\SHIT HAPPENING\Ruurd files\Extracted'
with zipfile.ZipFile(path_to_zip_file, 'r') as zip_ref:
    zip_ref.extractall(directory_to_extract_to)

# second extraction
path_to_zip_file = r'C:\Users\310295192\Desktop\SHIT HAPPENING\Ruurd files\Extracted\Complex report - missing rates.zip'
directory_to_extract_to = r'C:\Users\310295192\Desktop\Work\Rates\Dashboard\reports'
with zipfile.ZipFile(path_to_zip_file, 'r') as zip_ref:
    zip_ref.extractall(directory_to_extract_to)

#going to directory
os.chdir(r'C:\Users\310295192\Desktop\Work\Rates\Dashboard\reports')
        # going through each filename
for dirpath,dirnames,filenames in os.walk(os.curdir):
    for naming in filenames:
        # if there is a file with Ruurd name
        print(naming)
        if naming[:3] == "Mis":
            # it corrects the file to the name we need for easier processing
            os.rename(naming, naming_file)
            break

# removing unnecessary files
os.remove(r"C:\Users\310295192\Desktop\SHIT HAPPENING\Ruurd files\complex_report.zip")
os.remove(r"C:\Users\310295192\Desktop\SHIT HAPPENING\Ruurd files\Extracted\Complex report - missing rates.zip")

# reading data from csv. Separator is ;
dash = pd.read_csv(naming_file,sep = ';',dtype={'Port of loading':str,'Port of discharge':str,
                                                    'Container number':str,'Bill of lading':str,
                                                'Bookingnumber':str,'HAWB':str,
                                                'Leg 2 Carrier Name':str,'Leg 3 Carrier Name':str,
                                                'Container type':str,'Leg 2 Transport Mode':str,
                                                'Leg 3 Transport Mode':str})


#cleaning the data
# deleting some carriers
searchfor = ['Customer Pickup', 'PHILIPS CARRIER','Central Transport International','FedEx Freight PRIORITY',
                 'FedEx Freight ECONOMY','ZIPLINE LOGISTICS','UPS Express','TNT Special services']
dash = dash[~dash['Leg 1 Carrier Name'].str.contains('|'.join(searchfor),regex=True,na=False)]

# deleting PNAM shipments
filt_dash = dash['Shipment'].str.contains('PNAM',regex=True,na=False)
# creating dataframe for raod shipments
dash = dash[~filt_dash]

filt_dash_2 = dash['Shipment Booker Name'].isin(['Biazet LTL','Biazet FTL','Philips DC Venray',
                                                  'Philips DC Corbas','Philips DC Corby','Philips DC Batta'])

dash = dash[~filt_dash_2]
# dash.to_excel(r"C:\Users\310295192\Desktop\whatthefuck.xlsx", sheet_name='FCL rates', index=False)
# FUCkING FORMAT TO BE CHEckED
print(dash['Total costs Sales'])
filt_dash_3 = dash['Total costs Sales'] == 0
print(dash['Total costs Sales'])
dash = dash[filt_dash_3]
print(dash)
print('dash_3')
# switching names of domains to more readable format
dash['Domain Code'].replace({'PHILIPSDCN':'DC North',"PHILIPSIND":"India",
                    "PHILIPSDCS":"DC South",'PHILIPSDCE':"DC East",
                    'DAP':"CL","PHILIPSNAM":"NAM",
                    "PHILIPSR1":"SRC",'Domain PHILIPSM1':'MATC'},inplace=True)

# changing the transport mode for leg 2 to upper case letters
dash['Leg 2 Transport Mode'] = dash['Leg 2 Transport Mode'].map(lambda x: x.upper() if isinstance(x,str) else x)

# setting up filter for SEA shipments. Regex allows to look for both FCL and LCL with one string and "na" allows to keep empty values as faluse
filt_sea = dash['Leg 2 Transport Mode'].str.contains('SEA FCL',regex=True,na=False)
filt_sea_sub = dash['Leg 2 Carrier Name'].str.contains('sub',regex=True,na=False)
filt_lcl= dash['Leg 2 Carrier Name'].str.contains('LCL|Rail',regex=True,na=False)

# creating dataframe for sea shipments
dash_sea = dash[filt_sea]
dash_sea = dash_sea[~filt_sea_sub]
filt_sea_bill = dash_sea['Billable Indicator Leg 2']=='Yes'
dash_sea = dash_sea[filt_sea_bill]
dash_lcl =dash[filt_lcl]
filt_lcl_bill = dash_lcl['Billable Indicator Leg 2']=='Yes'
dash_lcl = dash_lcl[filt_lcl_bill]


# deleting Nippon from LCL
filt_dashboard = dash_lcl['Leg 2 Carrier Name'].str.contains('ppon|DHL LCL|mindo|DSV', regex=True, na=False)
dash_lcl = dash_lcl[~filt_dashboard]

# # adding expeditors lcl
#
# # creating the filter
# exp_lcl_filt = (dash['Leg 2 Carrier Name'].str.contains('Expeditors',regex=True,na=False) &
#                 dash['Leg 2 Transport Mode'].str.contains('LCL',regex=True,na=False))
#
# # applying the filter
# dash_exp_lcl = dash[exp_lcl_filt]
#
# # concatenating the tables
# dash_lcl = pd.concat([dash_exp_lcl,dash_lcl])

# setting up filter for AIR shipments
filt_air = dash['Leg 2 Transport Mode'].str.contains('SERVICE LEVEL',regex=True,na=False)
# creating dataframe for air shipments
dash_air = dash[filt_air]

# further filtering shipments
filt_air_bill =dash_air['Billable Indicator Leg 2']=='Yes'
dash_air = dash_air[filt_air_bill]



# setting up a filter for ROAD shipments
# filt_road = dash['Leg 2 Transport Mode'].isnull()
filt_road = dash['Leg 2 Transport Mode'].str.contains('FCL|LCL|LEVEL|DANGEROUS',regex=True,na=False)
# creating dataframe for raod shipments
dash_road = dash[~filt_road]

# creating dataframe for raod shipments
dash_road = dash[~filt_road]

dash_road = dash_road[~dash_road['Leg 1 Carrier Name'].str.contains('Unknown|UPS',regex=True,na=False)]

# filtering shipments for us
filt_road_us =  dash_road['Pickup country'].str.contains('US|MX|BR|CA',regex=True,na=False)
dash_road_us =  dash_road[filt_road_us]
dash_road_us =  dash_road[filt_road_us]

# deleting Unknown from US
filt_road_us_2 = dash_road_us['Leg 1 Carrier Name'].str.contains('Unknown|Fedex|Coast to Coast',
                                                          regex=True, na=False)
dash_road_us = dash_road_us[~filt_road_us_2]

# only billable
filt_road_us_2 = dash_road_us['Billable Indicator Leg 1'] == "Yes"
dash_road_us = dash_road_us[filt_road_us_2]


#filtering shipments for europe
filt_road_eu = dash_road['Pickup country'].str.contains('US|MX|BR|CA|CN|HK|IN',regex=True,na=False)
dash_road_eu = dash_road[~filt_road_eu]

# deleting shipments from unnecessary domains
filt_road_eu = dash_road_eu['Domain Code'].str.contains('DC North|India|DC South|DC East',regex=True,na=False)
dash_road_eu = dash_road_eu[~filt_road_eu]

# reading data from previous report
dashboard_file = r'C:\Users\310295192\Desktop\Work\Rates\Dashboard\rates dashboard {}.xlsx'.format(datum)

old_dashboard_sea = pd.read_excel(dashboard_file,
                                  sheet_name='FCL rates')
old_dashboard_lcl = pd.read_excel(dashboard_file,
                                  sheet_name='LCL rates')
old_dashboard_air = pd.read_excel(dashboard_file,
                                  sheet_name='AIR rates')
old_dashboard_road_eu = pd.read_excel(dashboard_file,
                                      sheet_name='EU ROAD rates')
old_dashboard_road_us = pd.read_excel(dashboard_file,
                                      sheet_name='US ROAD rates')


# list of dataframes
all_frames = {'dash_air':old_dashboard_air,
                       'dash_lcl':old_dashboard_lcl,
              'dash_road_eu':old_dashboard_road_eu,
              'dash_road_us':old_dashboard_road_us,
              'dash_sea':old_dashboard_sea}

# funtion to create hyperlink
def make_hyperlink(value,domain):
    url = "https://www.infodis.net/philips/net/Shipment/Details?InfodisNumber={}&DomainCode={}"
    return '=HYPERLINK("%s", "%s")' % (url.format(value,domain), value)

# looping over all dataframes and assiging new columns
for frame,old_frame in all_frames.items():
    # solving the isue with key element
    print(frame)
    if frame == 'dash_air':
        frame = dash_air
    elif frame == 'dash_lcl':
        frame = dash_lcl
    elif frame == 'dash_sea':
        frame = dash_sea
    elif frame == 'dash_road_eu':
        frame = dash_road_eu
    elif frame == 'dash_road_us':
        frame = dash_road_us


    frame['Action taken'] = np.nan
    frame['Pending/Solved'] = np.nan
    frame['Mail sent to IMS/Key-user'] = np.nan
    frame['Date of email'] = np.nan
    frame['Email subject'] = np.nan
    frame['Delay in answer'] = np.nan
    frame['Comments'] = np.nan

    # vlookup action taken
    frame['Action taken'] = frame['Shipment'].map(
        old_frame.set_index("Shipment")["Action taken"].to_dict())
    # vlook up email sent indicator, date of email, email subject, delay and comment
    frame['Pending/Solved'] = frame['Shipment'].map(
        old_frame.set_index("Shipment")["Pending/Solved"].to_dict())
    frame['Mail sent to IMS/Key-user'] = frame['Shipment'].map(
        old_frame.set_index("Shipment")['Mail sent to IMS/Key-user'].to_dict())
    frame['Date of email'] = frame['Shipment'].map(
        old_frame.set_index("Shipment")['Date of email'].to_dict())
    frame['Email subject'] = frame['Shipment'].map(
        old_frame.set_index("Shipment")['Email subject'].to_dict())
    # frame['Delay in answer'] = frame['Shipment'].map(old_dashboard_road_us.set_index("Shipment")['Delay in answer'].to_dict())
    frame['Comments'] = frame['Shipment'].map(
        old_frame.set_index("Shipment")["Comments"].to_dict())
    # assing column with today date
    frame['today'] = today

    # adjust data type for email date
    frame.loc[frame['Date of email'].notna(), 'Date of email'] = frame['Date of email'].dt.date

    # fill in na values and calculate dates
    frame['Delay in answer'] = frame['Delay in answer'].fillna(-1)
    frame['Delay in answer'] = (frame.loc[frame['Delay in answer'].notna(),
                                                        'today'] - frame.loc[
                                           frame['Delay in answer'].notna(),
                                           'Date of email']).dt.days.fillna(-1).astype(int)
    # drop today column
    frame.drop(columns=['today'], axis=1, inplace=True)

    # creating additional column for combining infodis number and domain
    frame['hyperlink'] = frame['Infodis'].astype(str) + frame['Domain Code'].apply(lambda x: 'DAP' if x == 'CL' else ("PHILIPSR1" if x=='SRC' else 'PHILIPSM1'))
    # switch infodis column to present hyperlink to infodis
    frame['Infodis'] = frame['hyperlink'].apply(
        lambda x: make_hyperlink(x[:-5], x[-3:]) if x[-1:] == 'P' else make_hyperlink(x[:-11], x[-9:]))
    # dropping helper column
    frame.drop(columns=['hyperlink'],axis=1,inplace=True)

    # change format for dates
    frame['Booking date'] = pd.to_datetime(frame['Booking date'],dayfirst=True).dt.date
    frame['Leg 1 Pickup Planned'] = pd.to_datetime(frame['Leg 1 Pickup Planned'],dayfirst=True).dt.date
    frame['Leg 1 Pickup Actual'] = pd.to_datetime(frame['Leg 1 Pickup Actual'],dayfirst=True).dt.date
    frame['Leg 2 Pickup Planned'] = pd.to_datetime(frame['Leg 2 Pickup Planned'],dayfirst=True).dt.date
    frame['Leg 2 Pickup Actual'] = pd.to_datetime(frame['Leg 2 Pickup Actual'],dayfirst=True).dt.date
    frame['Leg 3 Delivery Actual'] = pd.to_datetime(frame['Leg 3 Delivery Actual'],dayfirst=True).dt.date

    # sorting to see NAs first

    frame.sort_values(by='Pending/Solved',ascending=True,na_position='first',inplace=True)

    frame = None




# Create a Pandas Excel writer using XlsxWriter as the engine.

filename = r'C:\Users\310295192\Desktop\Work\Rates\Dashboard\rates dashboard {}.xlsx'.format(today)
writer = pd.ExcelWriter(filename,
                        engine='xlsxwriter', date_format='yyyy-mm-dd', datetime_format='yyyy-mm-dd')

# Convert the dataframe to an XlsxWriter Excel object.
dash_sea.to_excel(writer, sheet_name='FCL rates', index=False)
dash_lcl.to_excel(writer, sheet_name='LCL rates', index=False)
dash_air.to_excel(writer, sheet_name='AIR rates', index=False)
dash_road_eu.to_excel(writer, sheet_name='EU ROAD rates', index=False)
dash_road_us.to_excel(writer, sheet_name='US ROAD rates', index=False)

# Get the xlsxwriter workbook and worksheet objects.
workbook  = writer.book
worksheet = writer.sheets['FCL rates']


# Get the xlsxwriter workbook and worksheet objects.
workbook1  = writer.book
worksheet1 = writer.sheets['LCL rates']

# Get the xlsxwriter workbook and worksheet objects.
workbook2  = writer.book
worksheet2 = writer.sheets['AIR rates']

# Get the xlsxwriter workbook and worksheet objects.
workbook3 = writer.book
worksheet3 = writer.sheets['EU ROAD rates']

# Get the xlsxwriter workbook and worksheet objects.
workbook4  = writer.book
worksheet4 = writer.sheets['US ROAD rates']


# some formatting set-up
blueformat = workbook.add_format({'bg_color':'#AED6F1'})
darkformat = workbook.add_format({'bg_color':'#3498DB'})
dateformat = workbook.add_format({'num_format':'yyyy-mm-dd'})
rows = dash_air.shape[0]
columns = dash_air.shape[1]

# list of all worksheets
worksheets = [worksheet,worksheet1,worksheet2,worksheet3,worksheet4]

# loop over all worksheets
for excel in worksheets:
    #set width of the column
    excel.set_column('A:AW',21)
    #excel.set_column('S:S', 21, dateformat)
# save file
writer.save()

# open worksheet
wb = openpyxl.load_workbook(filename)


# dictionary of worksheets and dataframes
pairing = {'FCL rates':dash_sea,'LCL rates':dash_lcl,
           'AIR rates':dash_air,'EU ROAD rates':dash_road_eu,
           'US ROAD rates':dash_road_us}
# looping over worksheets
for work,framing in pairing.items():

    # solving the isue with key element and assigning worksheet
    if work == 'FCL rates':
        sheet = wb['FCL rates']
    elif work == 'LCL rates':
        sheet = wb['LCL rates']
    elif work == 'AIR rates':
        sheet = wb['AIR rates']
    elif work == 'EU ROAD rates':
        sheet = wb['EU ROAD rates']
    elif work == 'US ROAD rates':
        sheet = wb['US ROAD rates']


    # set up color for font
    ft = Font(color="fcfcfa")
    # table dimensions
    rows = framing.shape[0] + 1
    columns = framing.shape[1]
    # loop through each row
    for x in range(1, rows + 1):
        for y in range(1, columns + 1):
            # if case is solved then color it...
            if sheet.cell(row=x, column=44).value == 'solved':
                # green
                sheet.cell(row=x, column=y).fill = PatternFill(fgColor="407832", fill_type="solid")
                sheet.cell(row=x, column=y).font = ft
            # if case is not solved color it ....
            # orange
            elif sheet.cell(row=x, column=44).value == 'pending':
                sheet.cell(row=x, column=y).fill = PatternFill(fgColor="de9d23", fill_type="solid")
                sheet.cell(row=x, column=y).font = ft
    # save workbook with zoom setting 75
    sheet.sheet_view.zoomScale = 75


wb.save(filename)
wb.close()

print(datetime.datetime.now())