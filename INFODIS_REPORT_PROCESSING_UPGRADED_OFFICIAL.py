#concept mis to download all reports from today and update them based on last day


import win32com.client
import datetime
import os
import time
import openpyxl
from openpyxl.styles import PatternFill, Font,Color
import pandas as pd
import matplotlib.pyplot as plt
import win32print, win32api,win32com.client
import numpy as np

print(datetime.datetime.now())
#getting to inbox
folder = win32com.client.Dispatch("Outlook.Application").GetNameSpace("MAPI").GetDefaultFolder(6)
#getting to infodis folder
subfolder=folder.Folders(3)
#getting to all emails
email = subfolder.Items
#checkiing the number of emails to further take this value into the loop
x=len(email)
print(x)
#sort all emails that we have in the folder
email.Sort('ReceivedTime')
first=x-7
#starts the loop through the emails
for infodis in range(8):
#checks the number of email after sorting and extract data on it
    message = email.Item(first + infodis)
    bodyofemail = message.body
    subjectofemail=message.subject
    #checks the date of the email and assing the proper date
    when = message.SentOn
    when_formated= datetime.datetime.strftime(when,'%Y-%m-%d')
    #now we will be checking each email subject and save it in proper folder
    print(subjectofemail)
    if subjectofemail == "FCL missing rates":
        attachment = message.Attachments.Item(1)
        # if the email is for FCL missing rates - it saves it in respective folder
        attachment.SaveAsFile(r'C:\Users\310295192\Desktop\Work\Rates\Sea\reports\FCL report '+when_formated+'.xlsx')
    elif subjectofemail == "NAM missing rates overview":
        attachment = message.Attachments.Item(1)
        # if the email is for FCL missing rates - it saves it in respective folder
        attachment.SaveAsFile(
            r'C:\Users\310295192\Desktop\Work\Rates\Road\US\reports\report road NAM ' + when_formated + '.xlsx')
    elif subjectofemail == "Missing Rates AIR":
        attachment = message.Attachments.Item(1)
        # if the email is for AIR missing rates - it saves it in respective folder
        attachment.SaveAsFile(r'C:\Users\310295192\Desktop\Work\Rates\Air\reports\report AIR '+when_formated+'.xlsx')
    elif subjectofemail == "Missing Rates Road Europe -NL28 IT59 GB71 NL59 RO59":
        attachment = message.Attachments.Item(1)
        # if the email is for ROAD EU missing rates - it saves it in respective folder
        attachment.SaveAsFile(r'C:\Users\310295192\Desktop\Work\Rates\Road\reports\report road '+when_formated+'.xls')
    elif subjectofemail == "LCL missing rates for DB Schenker":
        attachment = message.Attachments.Item(1)
        # if the email is for LCL DB Schenker missing rates - it saves it in respective folder
        attachment.SaveAsFile(r'C:\Users\310295192\Desktop\Work\Rates\LCL\LCL reports\DB LCL '+when_formated+'.xlsx')
    elif subjectofemail == "No. of FCL shipments":
        attachment = message.Attachments.Item(1)
# if the email is for no. of shipments - it saves it in respective folder
        attachment.SaveAsFile(r'C:\Users\310295192\Desktop\Work\Rates\Daily management\No. of FCL shipments\FCL_shipments '+when_formated+'.xlsx')
    #comment
    elif subjectofemail == "EU missing rates ROAD":
        # if the email is for ROAD EU missing rates - it saves it in respective folder
        attachment = message.Attachments.Item(1)
        attachment.SaveAsFile(
            r'C:\Users\310295192\Desktop\Work\Rates\Road\reports complex\report road ' + when_formated + '.xlsx')
# NEW CODE WITH PANDAS FOR UPDATING FILES


# checking the name of the today
checking_date = datetime.datetime.today().strftime("%A")


# if it is not monday it takes one day before
if  checking_date == 'Monday':
    datum = datetime.datetime.today() - datetime.timedelta(3)
    datum = datum.strftime('%Y-%m-%d')
    # datum = '2019-08-18'
# if it is monday it takes three days before
else:
    datum = datetime.datetime.today() - datetime.timedelta(1)
    datum = datum.strftime('%Y-%m-%d')

today = datetime.date.today().strftime('%Y-%m-%d')

#ROAD
# turned off as i am not using it anymore
# # load two files - old one and new one
# old_road = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Road\reports\report road {} updated.xlsx'.format(datum))
# new_road = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Road\reports\report road {}.xls'.format(today))
#
# # from old one take only values that are below 0 cost
# old_road = old_road[old_road['Leg 1, Total costs']==0]
#
# # assign new columns to new report/dataframe
# new_road['Action taken']=np.nan
# new_road['Mail sent to IMS/key-user']=np.nan
# new_road['date of Email']=np.nan
# new_road['Email subject']=np.nan
# new_road['Delay']=np.nan
# new_road['Comments']=np.nan
#
# # leaving shipments with 0 costs only
# new_road = new_road[new_road['Leg 1, Total costs']==0].reset_index(drop=True)
#
# # vlookup action taken
# new_road['Action taken'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Action taken"].to_dict())
#
# # vlook up email sent indicator, date of email, email subject, delay and comment
# new_road['Mail sent to IMS/key-user'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Mail sent to IMS/key-user"].to_dict())
# new_road['date of Email'] = new_road['Infodis'].map(old_road.set_index("Infodis")["date of Email"].to_dict())
# new_road['Email subject'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Email subject"].to_dict())
# new_road['Delay'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Delay"].to_dict())
# new_road['Comments'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Comments"].to_dict())
#
# today = datetime.date.today()
#
# # assign column today for calculations
# new_road['today'] = today
#
# # adjust data type for email date
# new_road.loc[new_road['date of Email'].notna(),'date of Email'] = new_road['date of Email'].dt.date
#
#
# # fill na with -1
# new_road['Delay'] = new_road['Delay'].fillna(-1)
#
# # switch to integers
# new_road['Delay'] = new_road['Delay'].astype(int)
# # calculate the delay in days
# new_road.loc[new_road.Delay.notna(),'Delay'] = (new_road.loc[new_road.Delay.notna(),'today']- new_road.loc[new_road.Delay.notna(),'date of Email'])
# #new_road['Delay'] = new_road['Delay'].astype(str)
# new_road['Delay'] = new_road['Delay'].replace('-1', np.nan)
#
# # fill in na values and calculate dates
# new_road['Delay'] = new_road['Delay'].fillna(-1)
# new_road['Delay']=(new_road.loc[new_road.Delay.notna(),'today']- new_road.loc[new_road.Delay.notna(),'date of Email']).dt.days.fillna(-1).astype(int)
#
#
# # drop unnecessary columns
# new_road.drop(columns=['today','Shipment','Leg 1, Standard Freight','Leg 1, Fuel charge','Leg 1, Currency',
#                       'Invoice nr','Booking nr','CMR','Leg 2, Transport Mode',
#                       'Leg 2, Billable','Leg 2, Carrier Code','Leg 2, Carrier Name',
#                       'Booker id','Booker name','Shipper id','Consignee id',
#                       'Terms','Status','Leg 1, Pickup Actual','Leg 1, Pickup Planned',
#                       'Leg 1, Delivery Actual','Leg 1, Delivery Planned',
#                       'Leg 1, Leg delivery country','Leg n, Delivery Estimated',
#                       'Carrier Prebill number'], axis=1, inplace=True)
#
# # rename column
# new_road.rename(columns={"Shipment.1":"Shipment"},inplace=True)
#
# # create excel object in pandas with proper formatting
# # filename = r'C:\Users\310295192\Desktop\roadtest.xlsx'
# filename = r'C:\Users\310295192\Desktop\Work\Rates\Road\reports\report road {} updated.xlsx'.format(today)
# writer = pd.ExcelWriter(filename, engine='xlsxwriter',date_format = 'yyyy-mm-dd', datetime_format='yyyy-mm-dd')
#
# # dump data to excel
# new_road.to_excel(writer, sheet_name='Missing_Rates',index=False)
#
# # prepare worksheet for working and the formating
# workbook  = writer.book
# worksheet = writer.sheets['Missing_Rates']
# blueformat = workbook.add_format({'bg_color':'#AED6F1'})
# darkformat = workbook.add_format({'bg_color':'#3498DB'})
# rows = new_road.shape[0]
# columns = new_road.shape[1]
#
# #set width of the column
# worksheet.set_column('A:Z',20)
#
# # save file
# writer.save()
#
# # open worksheet
# wb = openpyxl.load_workbook(filename)
# # assign workbook and worksheet
# sheet = wb['Missing_Rates']
# # set up color for font
# ft = Font(color="fcfcfa")
# # table dimensions
# rows = new_road.shape[0] + 1
# columns = new_road.shape[1]
# # loop through each row
# for x in range(1,rows+1):
#     for y in range(1,columns+1):
#         # if case is solved then color it...
#         if sheet.cell(row=x,column=20).value == 'solved':
#         # green
#             sheet.cell(row=x,column=y).fill = PatternFill(fgColor="407832", fill_type = "solid")
#             sheet.cell(row=x,column=y).font = ft
#         # if case is not solved color it ....
#         # orange
#         elif (sheet.cell(row=x,column=20).value == 'pending') or (sheet.cell(row=x,column=20).value == 'yes') :
#             sheet.cell(row=x,column=y).fill = PatternFill(fgColor="de9d23", fill_type = "solid")
#             sheet.cell(row=x,column=y).font = ft
# # save workbook with zoom setting 75
# sheet.sheet_view.zoomScale = 75
# wb.save(filename)
# wb.close()


# ROAD EU

# checking the name of the today
checking_date = datetime.datetime.today().strftime("%A")


# if it is not monday it takes one day before
if  checking_date == 'Monday':
    datum = datetime.datetime.today() - datetime.timedelta(3)
    datum = datum.strftime('%Y-%m-%d')
# if it is monday it takes three days before
else:
    datum = datetime.datetime.today()
    datum = datetime.datetime.today() - datetime.timedelta(1)
    datum = datum.strftime('%Y-%m-%d')

today = datetime.date.today().strftime('%Y-%m-%d')


# load two files - old one and new one
old_road = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Road\reports complex\report road {} updated.xlsx'.format(datum))
new_road = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Road\reports complex\report road {}.xlsx'.format(today))

# from old one take only values that are below 0 cost
old_road = old_road[old_road['Leg 1, Total costs']==0]

# assign new columns to new report/dataframe
new_road['Action taken']=np.nan
new_road['Mail sent to IMS/key-user']=np.nan
new_road['date of Email']=np.nan
new_road['Email subject']=np.nan
new_road['Delay']=np.nan
new_road['Comments']=np.nan

# leaving shipments with 0 costs only
new_road = new_road[new_road['Leg 1, Total costs']==0].reset_index(drop=True)
new_road = new_road[~new_road['Booker name'].isin(['Biazet LTL','Biazet FTL','Philips DC Venray',
                                                  'Philips DC Corbas','Philips DC Corby','Philips DC Batta'])]


# vlookup action taken
new_road['Action taken'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Action taken"].to_dict())

# vlook up email sent indicator, date of email, email subject, delay and comment
new_road['Mail sent to IMS/key-user'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Mail sent to IMS/key-user"].to_dict())
new_road['date of Email'] = new_road['Infodis'].map(old_road.set_index("Infodis")["date of Email"].to_dict())
new_road['Email subject'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Email subject"].to_dict())
new_road['Delay'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Delay"].to_dict())
new_road['Comments'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Comments"].to_dict())

# check today date

today = datetime.date.today()

four_days = today - datetime.timedelta(4)

# assign column today for calculations
new_road['today'] = today

# adjust data type for email date
new_road.loc[new_road['date of Email'].notna(),'date of Email'] = new_road['date of Email'].dt.date

# fill na with -1
new_road['Delay'] = new_road['Delay'].fillna(-1)
#new_road.loc[new_road.Delay.notna(),'Delay'] = new_road.loc[new_road.Delay.notna(),'today']- new_road.loc[new_road.Delay.notna(),'date of Email']
# switch to integers
new_road['Delay'] = new_road['Delay'].astype(int)
# calculate the delay in days
new_road.loc[new_road.Delay.notna(),'Delay'] = (new_road.loc[new_road.Delay.notna(),'today']- new_road.loc[new_road.Delay.notna(),'date of Email'])
#new_road['Delay'] = new_road['Delay'].astype(str)
new_road['Delay'] = new_road['Delay'].replace('-1', np.nan)

# fill in na values and calculate dates
new_road['Delay'] = new_road['Delay'].fillna(-1)
new_road['Delay']=(new_road.loc[new_road.Delay.notna(),'today']- new_road.loc[new_road.Delay.notna(),'date of Email']).dt.days.fillna(-1).astype(int)

# drop unnecessary columns
new_road.drop(columns=['today','Shipment','Leg 1, Standard Freight','Leg 1, Fuel charge','Leg 1, Currency',
                      'Invoice nr','Booking nr','CMR','Leg 2, Transport Mode',
                      'Leg 2, Billable','Leg 2, Carrier Code','Leg 2, Carrier Name',
                      'Booker id','Booker name','Shipper id','Consignee id',
                      'Terms','Status','Leg 1, Pickup Actual','Leg 1, Pickup Planned',
                      'Leg 1, Delivery Actual','Leg 1, Delivery Planned',
                      'Leg 1, Leg delivery country','Leg n, Delivery Estimated',
                      'Carrier Prebill number'], axis=1, inplace=True)

# rename column
new_road.rename(columns={"Shipment.1":"Shipment"},inplace=True)


new_road['Action taken'] = np.where(new_road['Leg 1, Carrier Name'].str.contains("TNT Special services"),'solved',new_road['Action taken'])
new_road['Comments'] = np.where(new_road['Leg 1, Carrier Name'].str.contains("TNT Special services"),'Ad-hoc shipment (without rates)',new_road['Comments'])
# create excel object in pandas with proper formatting

filename = r'C:\Users\310295192\Desktop\Work\Rates\Road\reports complex\report road {} updated.xlsx'.format(today)
writer = pd.ExcelWriter(filename, engine='xlsxwriter',date_format = 'yyyy-mm-dd', datetime_format='yyyy-mm-dd')


# dump data to excel
new_road.to_excel(writer, sheet_name='Missing_Rates',index=False)

# prepare worksheet for working and the formating
workbook  = writer.book
worksheet = writer.sheets['Missing_Rates']
blueformat = workbook.add_format({'bg_color':'#AED6F1'})
darkformat = workbook.add_format({'bg_color':'#3498DB'})
rows = new_road.shape[0]
columns = new_road.shape[1]

#set width of the column
worksheet.set_column('A:Z',20)

# save file
writer.save()

# playing with openpyxl and formatting
import openpyxl
# import modules
from openpyxl.styles import PatternFill, Font,Color
# open worksheet
wb = openpyxl.load_workbook(filename)
# assign workbook and worksheet
sheet = wb['Missing_Rates']
# set up color for font
ft = Font(color="fcfcfa")
# table dimensions
rows = new_road.shape[0] + 1
columns = new_road.shape[1]
# loop through each row
for x in range(1,rows+1):
    for y in range(1,columns+1):
        # if case is solved then color it...
        if sheet.cell(row=x,column=20).value == 'solved':
        # green
            sheet.cell(row=x,column=y).fill = PatternFill(fgColor="407832", fill_type = "solid")
            sheet.cell(row=x,column=y).font = ft
        # if case is not solved color it ....
        # orange
        elif (sheet.cell(row=x,column=20).value == 'pending') or (sheet.cell(row=x,column=20).value == 'yes') :
            sheet.cell(row=x,column=y).fill = PatternFill(fgColor="de9d23", fill_type = "solid")
            sheet.cell(row=x,column=y).font = ft
# save workbook with zoom setting 75
sheet.sheet_view.zoomScale = 75
wb.save(filename)
wb.close()


# ROAD NAM

# checking the name of the today
checking_date = datetime.datetime.today().strftime("%A")


# if it is not monday it takes one day before
if  checking_date == 'Monday':
    datum = datetime.datetime.today() - datetime.timedelta(3)
    datum = datum.strftime('%Y-%m-%d')
# if it is monday it takes three days before
else:
    datum = datetime.datetime.today()
    datum = datetime.datetime.today() - datetime.timedelta(1)
    datum = datum.strftime('%Y-%m-%d')

today = datetime.date.today().strftime('%Y-%m-%d')

# load two files - old one and new one
old_road = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Road\US\reports\report road NAM {} updated.xlsx'.format(datum))
new_road = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Road\US\reports\report road NAM {}.xlsx'.format(today))

# from old one take only values that are below 0 cost
old_road = old_road[old_road['Leg 1, Total costs']==0]

old_road
# assign new columns to new report/dataframe
new_road['Action taken']=np.nan
new_road['Mail sent to IMS/key-user']=np.nan
new_road['date of Email']=np.nan
new_road['Email subject']=np.nan
new_road['Delay']=np.nan
new_road['Comments']=np.nan

# leaving shipments with 0 costs only
new_road = new_road[new_road['Leg 1, Total costs']==0].reset_index(drop=True)

# vlookup action taken
new_road['Action taken'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Action taken"].to_dict())

# vlook up email sent indicator, date of email, email subject, delay and comment
new_road['Mail sent to IMS/key-user'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Mail sent to IMS/key-user"].to_dict())
new_road['date of Email'] = new_road['Infodis'].map(old_road.set_index("Infodis")["date of Email"].to_dict())
new_road['Email subject'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Email subject"].to_dict())
new_road['Delay'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Delay"].to_dict())
new_road['Comments'] = new_road['Infodis'].map(old_road.set_index("Infodis")["Comments"].to_dict())

# check today date

today = datetime.date.today()

four_days = today - datetime.timedelta(4)

# assign column today for calculations
new_road['today'] = today

# adjust data type for email date
new_road.loc[new_road['date of Email'].notna(),'date of Email'] = new_road['date of Email'].dt.date

(new_road.loc[11,'today']-new_road.loc[11,'date of Email']).days

# fill na with -1
new_road['Delay'] = new_road['Delay'].fillna(-1)
#new_road.loc[new_road.Delay.notna(),'Delay'] = new_road.loc[new_road.Delay.notna(),'today']- new_road.loc[new_road.Delay.notna(),'date of Email']
# switch to integers
new_road['Delay'] = new_road['Delay'].astype(int)
# calculate the delay in days
new_road.loc[new_road.Delay.notna(),'Delay'] = (new_road.loc[new_road.Delay.notna(),'today']- new_road.loc[new_road.Delay.notna(),'date of Email'])
#new_road['Delay'] = new_road['Delay'].astype(str)
new_road['Delay'] = new_road['Delay'].replace('-1', np.nan)

#fill in na values and calculate dates
new_road['Delay'] = new_road['Delay'].fillna(-1)
new_road['Delay']=(new_road.loc[new_road.Delay.notna(),'today']- new_road.loc[new_road.Delay.notna(),'date of Email']).dt.days.fillna(-1).astype(int)

# drop unnecessary columns
new_road.drop(columns=['today','Shipment',
                      'Invoice nr','Booking nr','CMR','Leg 2, Transport Mode',
                      'Leg 2, Billable','Leg 2, Carrier Code','Leg 2, Carrier Name',
                      'Booker id','Shipper id','Consignee id',
                      'Terms','Status','Leg 1, Pickup Actual','Leg 1, Pickup Planned',
                      'Leg 1, Delivery Actual','Leg 1, Delivery Planned',
                      'Leg 1, Leg delivery country','Leg n, Delivery Estimated',
                      'Carrier Prebill number'], axis=1, inplace=True)

# rename column
new_road.rename(columns={"Shipment Duplicate(1)":"Shipment"},inplace=True)
new_road['Action taken'] = np.where(new_road.Shipment.str.contains("NAM"),'solved',new_road['Action taken'])
new_road['Comments'] = np.where(new_road.Shipment.str.contains("NAM"),'Not a case for PCT',new_road['Comments'])
# create excel object in pandas with proper formatting
# filename = r'C:\Users\310295192\Desktop\roadtest.xlsx'
filename = r'C:\Users\310295192\Desktop\Work\Rates\Road\US\reports\report road NAM {} updated.xlsx'.format(today)
writer = pd.ExcelWriter(filename, engine='xlsxwriter',date_format = 'yyyy-mm-dd', datetime_format='yyyy-mm-dd')


# dump data to excel
new_road.to_excel(writer, sheet_name='Missing_Rates_NAM',index=False)

# prepare worksheet for working and the formating
workbook  = writer.book
worksheet = writer.sheets['Missing_Rates_NAM']
blueformat = workbook.add_format({'bg_color':'#AED6F1'})
darkformat = workbook.add_format({'bg_color':'#3498DB'})
rows = new_road.shape[0]
columns = new_road.shape[1]

#set width of the column
worksheet.set_column('A:AD',20)

# save file
writer.save()

# playing with openpyxl and formatting
import openpyxl
# import modules
from openpyxl.styles import PatternFill, Font,Color
# open worksheet
wb = openpyxl.load_workbook(filename)
# assign workbook and worksheet
sheet = wb['Missing_Rates_NAM']
# set up color for font
ft = Font(color="fcfcfa")
# table dimensions
rows = new_road.shape[0] + 1
columns = new_road.shape[1]
print(rows,columns)
#loop through each row
for x in range(1,rows+1):
    for y in range(1,columns+1):
        # if case is solved then color it...
        if sheet.cell(row=x,column=24).value == 'solved':
        # green
            sheet.cell(row=x,column=y).fill = PatternFill(fgColor="407832", fill_type = "solid")
            sheet.cell(row=x,column=y).font = ft
        # if case is not solved color it ....
        # orange
        elif (sheet.cell(row=x,column=24).value == 'pending') or (sheet.cell(row=x,column=24).value == 'yes') :
            sheet.cell(row=x,column=y).fill = PatternFill(fgColor="de9d23", fill_type = "solid")
            sheet.cell(row=x,column=y).font = ft
# save workbook with zoom setting 75
sheet.sheet_view.zoomScale = 75
wb.save(filename)
wb.close()



# LCL

# checking the name of the today
checking_date = datetime.datetime.today().strftime("%A")


# if it is not monday it takes one day before
if  checking_date == 'Monday':
    datum = datetime.datetime.today() - datetime.timedelta(3)
    datum = datum.strftime('%Y-%m-%d')
    # datum = '2019-08-18'
# if it is monday it takes three days before
else:
    datum = datetime.datetime.today() - datetime.timedelta(1)
    datum = datum.strftime('%Y-%m-%d')
today = datetime.date.today().strftime('%Y-%m-%d')

# read files to dataframe

old_lcl = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\LCL\LCL reports\DB LCL {} updated.xlsx'.format(datum))
new_lcl = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\LCL\LCL reports\DB LCL {}.xlsx'.format(today))

# old_lcl.drop(columns=['Container number','Bill of lading','House B/L',
#        'Bookingnumber', 'GTNexus Id', 'Inttra ID','Leg 2, Transport confirmation', 'Confirmation status',
#        'Status Extended', 'Leg 2, Delivery Actual',
#        'Leg 3, Delivery Planned', 'Leg 3, Delivery Estimated',
#        'Leg 3, Delivery Actual', 'Leg 3, Delivery Actual-local',
#        'Shipment Bill Of Lading (BOL) attached by',
#        'Shipment Bill Of Lading (HBL) attached by', 'Invoice number',
#        'Booker Group', 'BL attached date', 'HBL attached date','Leg 2, Standard Freight Charges(purchase)',
#         'Billable Indicator Leg 1','Billable Indicator Leg 2', 'Billable Indicator Leg 3',
#         'Leg 1, Standard Freight Charges(purchase)',
#        'Leg 1, Total costs (purchase)', 'Leg 1, Pickup Actual', 'Container type(s)', 'Leg 2, Transport Mode',
#        'Weight', 'Volume', 'Actual volume', 'Carrier Prebill number',
#        'Leg 3, Standard Freight Charges(purchase)', 'Vessel name','Management fee (sales)',
#         'Cy fee (sales)', 'Booking agent fee (sales)',
#        'Leg 2, Full outgate - Actual', 'Leg 2, Container discharged',
#        'Notify id', 'Notify name (1)'],inplace=True)

# new_lcl.drop(columns=['Container number','Bill of lading','House B/L',
#        'Bookingnumber', 'GTNexus Id', 'Inttra ID','Leg 2, Transport confirmation', 'Confirmation status',
#        'Status Extended', 'Leg 2, Delivery Actual',
#        'Leg 3, Delivery Planned', 'Leg 3, Delivery Estimated',
#        'Leg 3, Delivery Actual', 'Leg 3, Delivery Actual-local',
#        'Shipment Bill Of Lading (BOL) attached by',
#        'Shipment Bill Of Lading (HBL) attached by', 'Invoice number',
#        'Booker Group', 'BL attached date', 'HBL attached date','Leg 2, Standard Freight Charges(purchase)',
#         'Billable Indicator Leg 1','Billable Indicator Leg 2', 'Billable Indicator Leg 3',
#         'Leg 1, Standard Freight Charges(purchase)',
#        'Leg 1, Total costs (purchase)', 'Leg 1, Pickup Actual', 'Container type(s)', 'Leg 2, Transport Mode',
#        'Weight', 'Volume', 'Actual volume', 'Carrier Prebill number',
#        'Leg 3, Standard Freight Charges(purchase)', 'Vessel name','Management fee (sales)',
#         'Cy fee (sales)', 'Booking agent fee (sales)',
#        'Leg 2, Full outgate - Actual', 'Leg 2, Container discharged',
#        'Notify id', 'Notify name (1)'],inplace=True)



new_lcl.drop(columns=['Booker id','Shipper id','Container number',
       'Bill of lading', 'Bookingnumber', 'GTNexus Id',
       'Shipment modifier/canceller',
       'Leg 2, Transport confirmation', 'Confirmation status',
       'Status Extended','Leg 2, Delivery Actual', 'Leg 3, Delivery Planned','Leg 3, Delivery Actual', 'Booker Group',
       'Shipment Bill Of Lading (BOL) attached by',
       'Shipment Bill Of Lading (HBL) attached by', 'Invoice number',
       'BL attached date', 'HBL attached date',
       'Leg 2, Total costs (purchase) EUR', 'Billable Indicator Leg 1',
       'Billable Indicator Leg 2', 'Billable Indicator Leg 3','Container type(s)',
        'Carrier Prebill number','Consignee id'],inplace=True)



# assign new columns to new report/dataframe
new_lcl['Action taken']=np.nan
new_lcl['Pending/Solved']=np.nan
new_lcl['Peter asked']=np.nan
new_lcl['Mail sent to IMS/Key-user']=np.nan
new_lcl['Date of email']=np.nan
new_lcl['Email subject']=np.nan
new_lcl['Delay in answer']=np.nan
new_lcl['Comments']=np.nan

# vlookup action taken
new_lcl['Action taken'] = new_lcl['Infodis'].map(old_lcl.set_index("Infodis")["Action taken"].to_dict())
# vlook up email sent indicator, date of email, email subject, delay and comment
new_lcl['Pending/Solved'] = new_lcl['Infodis'].map(old_lcl.set_index("Infodis")["Pending/Solved"].to_dict())
new_lcl['Peter asked'] = new_lcl['Infodis'].map(old_lcl.set_index("Infodis")["Peter asked"].to_dict())
new_lcl['Mail sent to IMS/Key-user'] = new_lcl['Infodis'].map(old_lcl.set_index("Infodis")["Mail sent to IMS/Key-user"].to_dict())
new_lcl['Date of email'] = new_lcl['Infodis'].map(old_lcl.set_index("Infodis")["Date of email"].to_dict())
new_lcl['Email subject'] = new_lcl['Infodis'].map(old_lcl.set_index("Infodis")["Email subject"].to_dict())
new_lcl['Delay in answer'] = new_lcl['Infodis'].map(old_lcl.set_index("Infodis")["Delay in answer"].to_dict())
new_lcl['Comments'] = new_lcl['Infodis'].map(old_lcl.set_index("Infodis")["Comments"].to_dict())

today = datetime.date.today()

# assign column today for calculations
new_lcl['today'] = today

# adjust data type for email date
new_lcl.loc[new_lcl['Date of email'].notna(),'Date of email'] = new_lcl['Date of email'].dt.date

# fill in na values and calculate dates
new_lcl['Delay in answer'] = new_lcl['Delay in answer'].fillna(-1)
new_lcl['Delay in answer']=(new_lcl.loc[new_lcl['Delay in answer'].notna(),'today']- new_lcl.loc[new_lcl['Delay in answer'].notna(),'Date of email']).dt.days.fillna(-1).astype(int)

# create excel object in pandas with proper formatting
# and by they way - drop today column
new_lcl.drop(columns=['today'],axis=1,inplace=True)
# filename = r'C:\Users\310295192\Desktop\lcltest.xlsx'
filename = r'C:\Users\310295192\Desktop\Work\Rates\LCL\LCL reports\DB LCL {} updated.xlsx'.format(today)
writer = pd.ExcelWriter(filename, engine='xlsxwriter',date_format = 'yyyy-mm-dd', datetime_format='yyyy-mm-dd')

# dump data to excel
new_lcl.to_excel(writer, sheet_name='Missing_Rates',index=False)

# # prepare worksheet for working and the formating
# workbook  = writer.book
# worksheet = writer.sheets['Missing_Rates']
# blueformat = workbook.add_format({'bg_color':'#AED6F1'})
# darkformat = workbook.add_format({'bg_color':'#3498DB'})
# rows = new_lcl.shape[0]
# columns = new_lcl.shape[1]
#
# #set width of the column
# worksheet.set_column('A:AF',20)
#
# # save file
# writer.save()
#
#
#
# # open worksheet
# wb = openpyxl.load_workbook(filename)
# # assign workbook and worksheet
# sheet = wb['Missing_Rates']
# # set up color for font
# ft = Font(color="fcfcfa")
# # table dimensions
# rows = new_lcl.shape[0] + 1
# columns = new_lcl.shape[1]
# # loop through each row
# for x in range(1,rows+1):
#     for y in range(1,columns+1):
#         # if case is solved then color it...
#         if sheet.cell(row=x,column=24).value == 'solved':
#         # green
#             sheet.cell(row=x,column=y).fill = PatternFill(fgColor="407832", fill_type = "solid")
#             sheet.cell(row=x,column=y).font = ft
#         # if case is not solved color it ....
#         # orange
#         elif sheet.cell(row=x,column=24).value == 'pending':
#             sheet.cell(row=x,column=y).fill = PatternFill(fgColor="de9d23", fill_type = "solid")
#             sheet.cell(row=x,column=y).font = ft
# # save workbook with zoom setting 75
# sheet.sheet_view.zoomScale = 75
# wb.save(filename)
# wb.close()


# prepare worksheet for working and the formating
workbook  = writer.book
worksheet = writer.sheets['Missing_Rates']
blueformat = workbook.add_format({'bg_color':'#AED6F1'})
darkformat = workbook.add_format({'bg_color':'#3498DB'})
rows = new_lcl.shape[0]
columns = new_lcl.shape[1]

#set width of the column
worksheet.set_column('A:AF',21)

# save file
writer.save()

# open worksheet
wb = openpyxl.load_workbook(filename)
# assign workbook and worksheet
sheet = wb['Missing_Rates']
# set up color for font
ft = Font(color="fcfcfa")
# table dimensions
rows = new_lcl.shape[0] + 1
columns = new_lcl.shape[1]
# loop through each row
for x in range(1,rows+1):
    for y in range(1,columns+1):
        # if case is solved then color it...
        if sheet.cell(row=x,column=21).value == 'solved':
        # green
            sheet.cell(row=x,column=y).fill = PatternFill(fgColor="407832", fill_type = "solid")
            sheet.cell(row=x,column=y).font = ft
        # if case is not solved color it ....
        # orange
        elif sheet.cell(row=x,column=21).value == 'pending':
            sheet.cell(row=x,column=y).fill = PatternFill(fgColor="de9d23", fill_type = "solid")
            sheet.cell(row=x,column=y).font = ft
# save workbook with zoom setting 75
sheet.sheet_view.zoomScale = 75
wb.save(filename)
wb.close()













#AIR

import datetime
# checking the name of the today
checking_date = datetime.datetime.today().strftime("%A")


# if it is not monday it takes one day before
if  checking_date == 'Monday':
    datum = datetime.datetime.today() - datetime.timedelta(3)
    datum = datum.strftime('%Y-%m-%d')
    # datum = '2019-08-18'
# if it is monday it takes three days before
else:
    datum = datetime.datetime.today() - datetime.timedelta(1)
    datum = datum.strftime('%Y-%m-%d')

today = datetime.date.today().strftime('%Y-%m-%d')

old_air = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Air\reports\report AIR {} updated.xlsx'.format(datum))
new_air = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Air\reports\report AIR {}.xlsx'.format(today))

#  dropping unnecessary columns
# old_air.drop(columns=['Booker id','Shipper id','Container number',
#        'Bill of lading', 'Bookingnumber', 'GTNexus Id',
#        'Shipment modifier/canceller', 'Booking date',
#        'Leg 2, Transport confirmation', 'Confirmation status',
#        'Status Extended', 'Leg 2, Pickup Planned','Leg 3, Delivery Actual', 'Booker Group',
#        'Shipment Bill Of Lading (BOL) attached by',
#        'Shipment Bill Of Lading (HBL) attached by', 'Invoice number',
#        'BL attached date', 'HBL attached date',
#        'Leg 2, Total costs (purchase) EUR', 'Billable Indicator Leg 1',
#        'Billable Indicator Leg 2', 'Billable Indicator Leg 3','Container type(s)',
#         'Carrier Prebill number','Consignee id'],inplace=True)

new_air.drop(columns=['Booker id','Shipper id','Container number',
       'Bill of lading', 'Bookingnumber', 'GTNexus Id',
       'Shipment modifier/canceller',
       'Leg 2, Transport confirmation', 'Confirmation status',
       'Status Extended','Leg 2, Delivery Actual', 'Leg 3, Delivery Planned','Leg 3, Delivery Actual', 'Booker Group',
       'Shipment Bill Of Lading (BOL) attached by',
       'Shipment Bill Of Lading (HBL) attached by', 'Invoice number',
       'BL attached date', 'HBL attached date',
       'Leg 2, Total costs (purchase) EUR', 'Billable Indicator Leg 1',
       'Billable Indicator Leg 2', 'Billable Indicator Leg 3','Container type(s)',
        'Carrier Prebill number','Consignee id'],inplace=True)

# assign new columns to new report/dataframe
new_air['Action taken']=np.nan
new_air['Pending/Solved']=np.nan
new_air['date of email']=np.nan
new_air['E-mail subject']=np.nan
new_air['Delay in answer']=np.nan
new_air['Comments from carrier']=np.nan
new_air['Comments']=np.nan

# vlookup action taken
new_air['Action taken'] = new_air['Infodis'].map(old_air.set_index("Infodis")["Action taken"].to_dict())
# vlook up email sent indicator, date of email, email subject, delay and comment
new_air['Pending/Solved'] = new_air['Infodis'].map(old_air.set_index("Infodis")["Pending/Solved"].to_dict())
new_air['date of email'] = new_air['Infodis'].map(old_air.set_index("Infodis")['date of email'].to_dict())
new_air['E-mail subject'] = new_air['Infodis'].map(old_air.set_index("Infodis")['E-mail subject'].to_dict())
new_air['Delay in answer'] = new_air['Infodis'].map(old_air.set_index("Infodis")['Delay in answer'].to_dict())
new_air['Comments from carrier'] = new_air['Infodis'].map(old_air.set_index("Infodis")['Comments from carrier'].to_dict())
new_air['Comments'] = new_air['Infodis'].map(old_air.set_index("Infodis")["Comments"].to_dict())

# check today date
today = datetime.date.today()

new_air['today'] = today
# adjust data type for email date
new_air.loc[new_air['date of email'].notna(),'date of email'] = new_air['date of email'].dt.date

# fill in na values and calculate dates
new_air['Delay in answer'] = new_air['Delay in answer'].fillna(-1)
new_air['Delay in answer']=(new_air.loc[new_air['Delay in answer'].notna(),'today']- new_air.loc[new_air['Delay in answer'].notna(),'date of email']).dt.days.fillna(-1).astype(int)

# create excel object in pandas with proper formatting
# and by they way - drop today column
new_air.drop(columns=['today'],axis=1,inplace=True)
# filename = r'C:\Users\310295192\Desktop\airtest.xlsx'
filename = r'C:\Users\310295192\Desktop\Work\Rates\Air\reports\report AIR {} updated.xlsx'.format(today)
writer = pd.ExcelWriter(filename, engine='xlsxwriter',date_format = 'yyyy-mm-dd', datetime_format='yyyy-mm-dd')

# dump data to excel
new_air.to_excel(writer, sheet_name='Missing_Rates',index=False)

# prepare worksheet for working and the formating
workbook  = writer.book
worksheet = writer.sheets['Missing_Rates']
blueformat = workbook.add_format({'bg_color':'#AED6F1'})
darkformat = workbook.add_format({'bg_color':'#3498DB'})
rows = new_air.shape[0]
columns = new_air.shape[1]

#set width of the column
worksheet.set_column('A:AF',21)

# save file
writer.save()

# open worksheet
wb = openpyxl.load_workbook(filename)
# assign workbook and worksheet
sheet = wb['Missing_Rates']
# set up color for font
ft = Font(color="fcfcfa")
# table dimensions
rows = new_air.shape[0] + 1
columns = new_air.shape[1]
# loop through each row
for x in range(1,rows+1):
    for y in range(1,columns+1):
        # if case is solved then color it...
        if sheet.cell(row=x,column=21).value == 'solved':
        # green
            sheet.cell(row=x,column=y).fill = PatternFill(fgColor="407832", fill_type = "solid")
            sheet.cell(row=x,column=y).font = ft
        # if case is not solved color it ....
        # orange
        elif sheet.cell(row=x,column=21).value == 'pending':
            sheet.cell(row=x,column=y).fill = PatternFill(fgColor="de9d23", fill_type = "solid")
            sheet.cell(row=x,column=y).font = ft
# save workbook with zoom setting 75
sheet.sheet_view.zoomScale = 75
wb.save(filename)
wb.close()


# SEA/FCL

# checking the name of the today
checking_date = datetime.datetime.today().strftime("%A")


# if it is not monday it takes one day before
if  checking_date == 'Monday':
    datum = datetime.datetime.today() - datetime.timedelta(3)
    datum = datum.strftime('%Y-%m-%d')
    # datum = '2019-08-18'
# if it is monday it takes three days before
else:
    datum = datetime.datetime.today() - datetime.timedelta(1)
    datum = datum.strftime('%Y-%m-%d')
today = datetime.date.today().strftime('%Y-%m-%d')

# read file to dataframes

# old_sea = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Sea\reports\FCL report 2019-08-16.xlsx',parse_dates=['Date of email'])
# new_sea = pd.read_excel(r'C:\Users\310295192\Desktop\FCL report 2019-08-17.xlsx')

old_sea = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Sea\reports\FCL report {} updated.xlsx'.format(datum),parse_dates=['Date of email'])
new_sea = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Sea\reports\FCL report {}.xlsx'.format(today))


# #  dropping unnecessary columns
# old_sea.drop(columns=['Consignee id','Original shipper name (1)','House B/L',
#                        'Bookingnumber', 'GTNexus Id', 'Inttra ID','Booking date (time)',
#                       'Leg 2, Transport confirmation (date)','Leg 2, Transport confirmation (time)',
#                       'Status Extended','Leg 2, Pickup Planned (time)', 'Leg 2, Pickup Actual (date)',
#                        'Leg 2, Pickup Actual (time)', 'Leg 2, Delivery Actual (date)',
#                        'Leg 2, Delivery Actual (time)','Leg 3, Delivery Planned (date)', 'Leg 3, Delivery Planned (time)',
#        'Leg 3, Delivery Estimated (date)', 'Leg 3, Delivery Estimated (time)',
#        'Leg 3, Delivery Actual (date)', 'Leg 3, Delivery Actual (time)',
#        'Leg 3, Delivery Actual-local (date)',
#        'Leg 3, Delivery Actual-local (time)',
#        'Shipment Bill Of Lading (BOL) attached by',
#        'Shipment Bill Of Lading (HBL) attached by', 'Invoice number',
#        'Booker Group', 'BL attached date (date)', 'BL attached date (time)',
#        'HBL attached date (date)', 'HBL attached date (time)',
#        'Leg 2, Standard Freight Charges(purchase)', 'Billable Indicator Leg 1',
#        'Billable Indicator Leg 2', 'Billable Indicator Leg 3','Leg 1, Standard Freight Charges(purchase)',
#        'Leg 1, Total costs (purchase)', 'Leg 1, Pickup Actual (date)',
#        'Leg 1, Pickup Actual (time)','Weight','Leg 2, Full outgate - Actual (date)',
#        'Leg 2, Full outgate - Actual (time)',
#        'Leg 2, Container discharged (date)',
#        'Leg 2, Container discharged (time)', 'Notify id', 'Notify name (1)',
#        'Notify address (1)',
#                      ],inplace=True)

# new_sea.drop(columns=['Consignee id','Original shipper name (1)','House B/L',
#                        'Bookingnumber', 'GTNexus Id', 'Inttra ID','Booking date (time)',
#                       'Leg 2, Transport confirmation (date)','Leg 2, Transport confirmation (time)',
#                       'Status Extended','Leg 2, Pickup Planned (time)', 'Leg 2, Pickup Actual (date)',
#                        'Leg 2, Pickup Actual (time)', 'Leg 2, Delivery Actual (date)',
#                        'Leg 2, Delivery Actual (time)','Leg 3, Delivery Planned (date)', 'Leg 3, Delivery Planned (time)',
#        'Leg 3, Delivery Estimated (date)', 'Leg 3, Delivery Estimated (time)',
#        'Leg 3, Delivery Actual (date)', 'Leg 3, Delivery Actual (time)',
#        'Leg 3, Delivery Actual-local (date)',
#        'Leg 3, Delivery Actual-local (time)',
#        'Shipment Bill Of Lading (BOL) attached by',
#        'Shipment Bill Of Lading (HBL) attached by', 'Invoice number',
#        'Booker Group', 'BL attached date (date)', 'BL attached date (time)',
#        'HBL attached date (date)', 'HBL attached date (time)',
#        'Leg 2, Standard Freight Charges(purchase)', 'Billable Indicator Leg 1',
#        'Billable Indicator Leg 2', 'Billable Indicator Leg 3','Leg 1, Standard Freight Charges(purchase)',
#        'Leg 1, Total costs (purchase)', 'Leg 1, Pickup Actual (date)',
#        'Leg 1, Pickup Actual (time)','Weight','Leg 2, Full outgate - Actual (date)',
#        'Leg 2, Full outgate - Actual (time)',
#        'Leg 2, Container discharged (date)',
#        'Leg 2, Container discharged (time)', 'Notify id', 'Notify name (1)',
#        'Notify address (1)',],inplace=True)


new_sea.drop(columns=['Booker id','Shipper id','Container number',
       'Bill of lading', 'Bookingnumber', 'GTNexus Id',
       'Shipment modifier/canceller',
       'Leg 2, Transport confirmation', 'Confirmation status',
       'Status Extended','Leg 2, Delivery Actual', 'Leg 3, Delivery Planned','Leg 3, Delivery Actual', 'Booker Group',
       'Shipment Bill Of Lading (BOL) attached by',
       'Shipment Bill Of Lading (HBL) attached by', 'Invoice number',
       'BL attached date', 'HBL attached date',
       'Leg 2, Total costs (purchase) EUR', 'Billable Indicator Leg 1',
       'Billable Indicator Leg 2', 'Billable Indicator Leg 3','Container type(s)',
        'Carrier Prebill number','Consignee id'],inplace=True)



# assign new columns to new report/dataframe
new_sea['Action taken']=np.nan
new_sea['Pending/Solved']=np.nan
new_sea['Peter asked']=np.nan
new_sea['Mail sent to IMS/Key-user']=np.nan
new_sea['Date of email']=np.nan
new_sea['Email subject']=np.nan
new_sea['Delay in answer']=np.nan
new_sea['Comments']=np.nan

# vlookup action taken
new_sea['Action taken'] = new_sea['Infodis'].map(old_sea.set_index("Infodis")["Action taken"].to_dict())
# vlook up email sent indicator, date of email, email subject, delay and comment
new_sea['Pending/Solved'] = new_sea['Infodis'].map(old_sea.set_index("Infodis")["Pending/Solved"].to_dict())
new_sea['Peter asked'] = new_sea['Infodis'].map(old_sea.set_index("Infodis")['Peter asked'].to_dict())
new_sea['Mail sent to IMS/Key-user'] = new_sea['Infodis'].map(old_sea.set_index("Infodis")['Mail sent to IMS/Key-user'].to_dict())
new_sea['Date of email'] = new_sea['Infodis'].map(old_sea.set_index("Infodis")['Date of email'].to_dict())
new_sea['Email subject'] = new_sea['Infodis'].map(old_sea.set_index("Infodis")['Email subject'].to_dict())
new_sea['Delay in answer'] = new_sea['Infodis'].map(old_sea.set_index("Infodis")["Delay in answer"].to_dict())
new_sea['Comments'] = new_sea['Infodis'].map(old_sea.set_index("Infodis")["Comments"].to_dict())


# check today date
today = datetime.date.today()

# assign column today for calculations
new_sea['today'] = today

# adjust data type for email date
new_sea.loc[new_sea['Date of email'].notna(),'Date of email'] = new_sea['Date of email'].dt.date

# fill in na values and calculate dates
new_sea['Delay in answer'] = new_sea['Delay in answer'].fillna(-1)
new_sea['Delay in answer']=(new_sea.loc[new_sea['Delay in answer'].notna(),'today']- new_sea.loc[new_sea['Delay in answer'].notna(),'Date of email']).dt.days.fillna(-1).astype(int)

new_sea.drop(columns=['today'],axis=1,inplace=True)

# create excel object in pandas with proper formatting
# and by they way - drop today column
# new_sea.drop(columns=['today','Actual volume','Carrier Prebill number','Leg 3, Standard Freight Charges(purchase)',
#                      'Vessel name','Management fee (sales)','Cy fee (sales)',
#                      'Booking agent fee (sales)','Volume'],axis=1,inplace=True)
# filename = r'C:\Users\310295192\Desktop\fcltest.xlsx'
filename= r'C:\Users\310295192\Desktop\Work\Rates\Sea\reports\FCL report {} updated.xlsx'.format(today)
writer = pd.ExcelWriter(filename, engine='xlsxwriter',date_format = 'yyyy-mm-dd', datetime_format='yyyy-mm-dd')

# dump data to excel
new_sea.to_excel(writer, sheet_name='Missing_Rates',index=False)

# # prepare worksheet for working and the formating
# workbook  = writer.book
# worksheet = writer.sheets['Missing_Rates']
# blueformat = workbook.add_format({'bg_color':'#AED6F1'})
# darkformat = workbook.add_format({'bg_color':'#3498DB'})
# rows = new_sea.shape[0]
# columns = new_sea.shape[1]
#
# #set width of the column
# worksheet.set_column('A:AF',20)
#
# # save file
# writer.save()
#
#
# # open worksheet
# wb = openpyxl.load_workbook(filename)
# # assign workbook and worksheet
# sheet = wb['Missing_Rates']
# # set up color for font
# ft = Font(color="fcfcfa")
# # table dimensions
# rows = new_sea.shape[0] + 1
# columns = new_sea.shape[1]
# # loop through each row
# for x in range(1,rows+1):
#     for y in range(1,columns+1):
#         # if case is solved then color it...
#         if sheet.cell(row=x,column=25).value == 'solved':
#         # green
#             sheet.cell(row=x,column=y).fill = PatternFill(fgColor="407832", fill_type = "solid")
#             sheet.cell(row=x,column=y).font = ft
#         # if case is not solved color it ....
#         # orange
#         elif sheet.cell(row=x,column=25).value == 'pending':
#             sheet.cell(row=x,column=y).fill = PatternFill(fgColor="de9d23", fill_type = "solid")
#             sheet.cell(row=x,column=y).font = ft
# # save workbook with zoom setting 75
# sheet.sheet_view.zoomScale = 75
# wb.save(filename)
# wb.close()


# prepare worksheet for working and the formating
workbook  = writer.book
worksheet = writer.sheets['Missing_Rates']
blueformat = workbook.add_format({'bg_color':'#AED6F1'})
darkformat = workbook.add_format({'bg_color':'#3498DB'})
rows = new_sea.shape[0]
columns = new_sea.shape[1]

#set width of the column
worksheet.set_column('A:AF',21)

# save file
writer.save()

# open worksheet
wb = openpyxl.load_workbook(filename)
# assign workbook and worksheet
sheet = wb['Missing_Rates']
# set up color for font
ft = Font(color="fcfcfa")
# table dimensions
rows = new_sea.shape[0] + 1
columns = new_sea.shape[1]
# loop through each row
for x in range(1,rows+1):
    for y in range(1,columns+1):
        # if case is solved then color it...
        if sheet.cell(row=x,column=21).value == 'solved':
        # green
            sheet.cell(row=x,column=y).fill = PatternFill(fgColor="407832", fill_type = "solid")
            sheet.cell(row=x,column=y).font = ft
        # if case is not solved color it ....
        # orange
        elif sheet.cell(row=x,column=21).value == 'pending':
            sheet.cell(row=x,column=y).fill = PatternFill(fgColor="de9d23", fill_type = "solid")
            sheet.cell(row=x,column=y).font = ft
# save workbook with zoom setting 75
sheet.sheet_view.zoomScale = 75
wb.save(filename)
wb.close()





















# #updating KPI report (code below)
# #Open file with KPI
# kpi = openpyxl.load_workbook(r"C:\Users\310295192\Desktop\Work\Rates\Daily management\daily management.xlsx")
#
#
# #Go to 'data' sheet
# sheet=kpi['data']
# #iterate through all cells to find the first empty cell
# for cell in sheet["E"]:
#     if cell.value==None:
#         #not first empty cell
#         lastrow = cell.row
#         break
# #go to file with # of shipments
#
# noship=openpyxl.load_workbook(r'C:\Users\310295192\Desktop\Work\Rates\Daily management\No. of FCL shipments\FCL_shipments '+when_formated+'.xlsx')
#
# sheetnoship=noship['data']
# #defines the number of shipments
# quantity = sheetnoship.max_row-1
#
# #open the file for missing rates
# missing_rates=openpyxl.load_workbook(r'C:\Users\310295192\Desktop\Work\Rates\Sea\reports\FCL report '+when_formated+'.xlsx')
#
#
# missing_rates_sh=missing_rates['data']
# #defines the number of missing rates
# missing_number = missing_rates_sh.max_row-1
#
# #write the values to KPI workbook (first missing rates, second number of shipments)
# sheet.cell(row=lastrow,column=5).value = missing_number
# sheet.cell(row=lastrow,column=6).value = quantity
#
# #closes all three workbooks
# kpi.save(filename=r"C:\Users\310295192\Desktop\Work\Rates\Daily management\daily management.xlsx")
# kpi.close()
# noship.close()
# missing_rates.close()

# if it is not monday it takes one day before
checking_date = datetime.datetime.today().strftime("%A")

# # if it is not monday it takes one day before
# if checking_date == 'Wednesday':
#     fname = r'C:\Users\310295192\Desktop\Work\Rates\Daily management\daily management.xlsx'
#     df = pd.read_excel(fname, sheet_name="data")
#     today = datetime.datetime.today()
#     two_weeks = today - datetime.timedelta(21)
#     # extracting the proper dataframe from file
#     fcl_rates = df[(df['Date'] <= today) & (df['Date'] >= two_weeks)]
#     fcl_rates['Date'] = pd.to_datetime(fcl_rates["Date"].astype(str), format='%Y-%m-%d')
#     # getting the date format and proper indexing
#     fcl_rates['Date'] = fcl_rates['Date'].dt.date
#     fcl_rates.index = fcl_rates['Date']
#     # multiplying values by 100 to have percentages
#     fcl_rates["Target"] = fcl_rates['Target'] * 100
#     fcl_rates['Shipments with rates'] = fcl_rates['Shipments with rates'] * 100
#     fcl_rates['Coverage'] = ((fcl_rates['All Shipments'] - fcl_rates['Shipments without rates']) / fcl_rates[
#         'All Shipments']) * 100
#     # assigning date table for x axis
#     dates = [x.strftime('%Y-%m-%d') for x in fcl_rates['Date']]
#     # assigning table with values for y axis
#     values = [round(x, 3) for x in fcl_rates['Coverage']]
#     # assigning values for target
#     target = [99.9 for x in range(len(fcl_rates['Shipments with rates']))]
#     coloring = []
#     # checking the color for bars, if good then green and if not - red
#     for x in values:
#         if x > 99.9:
#             coloring.append('#038518')
#         else:
#             coloring.append('#f5424b')
#     # setting up figure size
#     plt.figure(dpi=1024, figsize=(10, 6))
#     # creating bar chart
#     plt.bar(dates, values, color=coloring)
#     # creating line chart
#     plt.plot(dates, target, color='#031f85', linewidth=3, linestyle='--')
#     # rotating ticks
#     plt.xticks(dates, rotation='vertical')
#     # adding title
#     plt.title('FCL shipments with rates in infodis', fontsize=14)
#     # setting up label for y axis
#     plt.ylabel('%', fontsize=14, rotation='horizontal')
#     # setting up legend and location
#     plt.legend(["Target"], loc='upper center', bbox_to_anchor=(0.5, -0.25))
#     # "Packing" the figure
#     plt.tight_layout()
#     # setting up date for file name
#     x = datetime.datetime.today().strftime("%Y-%m-%d")
#     # creating file name
#     f_name = r'C:\Users\310295192\Desktop\Work\Rates\Daily management\DM_Graphs\DM_FCL {}.png'.format(x)
#     # saving figure
#     plt.savefig(f_name)
#     # printing figure
#     currentprinter = win32print.GetDefaultPrinter()
#     win32api.ShellExecute(0, "print", f_name, currentprinter, ".", 0)
#     os.startfile(r'C:\Users\310295192\Desktop\Work\Rates\Daily management\DM_Graphs')

# preparing missing rates for Panalpina

# checking the name of the today
checking_date = datetime.datetime.today().strftime("%A")
datum = ""
# if it is not monday it takes one day before
if checking_date == 'Monday':
    datum = datetime.datetime.today()
    datum = datum.strftime('%Y-%m-%d')
# creating variable for location
    location = r'C:\Users\310295192\Desktop\Work\Rates\Air\reports\report AIR {} updated.xlsx'.format(datum)
    # create dataframe
    df = pd.read_excel(location)
    # Indicated shipments with Panalpina as carrier
    df = df[df['Leg 2, Carrier Name'] == 'Panalpina AIR S']
    # creating only shipments that are not picked up anyhow
    report = df[df['Pending/Solved'].isna()]
    # cut columns
    # columing = [x for x in range(0, 13)] + [x for x in range(22, 28)] + [38, 39,42]
    columing = [x for x in range(0, 19)]
    print(columing)
    panalpina_air = report.iloc[0:, columing]
    # change the format of dates
    panalpina_air['Leg 2, Pickup Planned'].dt.strftime('%Y-%m-%d')
    panalpina_air['Leg 2, Pickup Actual'].dt.strftime('%Y-%m-%d')
    # panalpina_air['Leg 2, Delivery Actual'].dt.strftime('%Y-%m-%d')
    # panalpina_air['Leg 3, Delivery Planned'].dt.strftime('%Y-%m-%d')
    # panalpina_air['Leg 3, Delivery Actual'].dt.strftime('%Y-%m-%d')
    # create excel object in pandas with proper formatting
    filename = r'C:\Users\310295192\Desktop\Work\Rates\Air\Panalpina\Sent reports on missing lanes - Panalpina\Panalpina AIR missing rates {}.xlsx'.format(datum)
    writer = pd.ExcelWriter(filename, engine='xlsxwriter', date_format='yyyy-mm-dd', datetime_format='yyyy-mm-dd')
    # dump data to excel
    panalpina_air.to_excel(writer, sheet_name='Missing_Rates', index=False)
    # prepare worksheet for working and the formating
    workbook = writer.book
    worksheet = writer.sheets['Missing_Rates']
    blueformat = workbook.add_format({'bg_color': '#2a5fb5'})
    darkformat = workbook.add_format({'bg_color': '#012052'})
    rows = panalpina_air.shape[0]
    columns = panalpina_air.shape[1]
    # set width of the column
    worksheet.set_column('A:V', 20)
    # save file
    writer.save()


    # open worksheet
    wb = openpyxl.load_workbook(filename)
    # assign workbook and worksheet
    sheet = wb['Missing_Rates']
    # set up color for font
    ft = Font(color="FFFFFFFF")
    # table dimensions
    rows_pan = panalpina_air.shape[0] + 1
    columns_pan = panalpina_air.shape[1]
    # loop through each row
    for x in range(1, rows_pan + 1):
        for y in range(1, columns_pan + 1):
            if x % 2 == 0:
                # bright blue
                sheet.cell(row=x, column=y).fill = PatternFill(fgColor="83C8F7", fill_type="solid")
                sheet.cell(row=x, column=y).font = ft
            # dark blue
            else:
                sheet.cell(row=x, column=y).fill = PatternFill(fgColor="1192E8", fill_type="solid")
                sheet.cell(row=x, column=y).font = ft
    # save workbook
    wb.save(filename)
    wb.close()





# preparing email for what has been added and what was wrong
    outlook = win32com.client.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    #mail.To = 'DL_PCTPolandOptilo@philips.com'
    mail.To = 'GBSCTPhilips.SMB@panalpina.com; philips.prebilling@philips.com'
    mail.CC = 'peter.van.dijk@philips.com ; Herman.Houweling@panalpina.com'
    mail.Subject = 'Panalpina AIR Missing Rates ' + str(datetime.datetime.today().strftime("%Y-%m-%d"))
    mail.Body = 'Test'
    mail.HTMLBody = '<p>Dear Panalpina,</p><p>Please find attached the overview of missing rates in infodis.</p>'\
                    '<br><p>Kind Regards,</p><p>Maciej Janowski</p>' \
                    '<p>Junior Transport Specialist</p><p>Philips Polska Sp. z o.o </p>' \
                    '<p>PCT Poland</p>'
    # attaching log file
    attachment  = filename
    mail.Attachments.Add(attachment)
    # sending email
    mail.Save()



    # # procedure for F&D
    # # setting up today date
    # today = datetime.datetime.today().strftime("%Y-%m-%d")
    #
    # # name of the file
    # x = r'C:\Users\310295192\Desktop\Work\Rates\Road\reports\report road {} updated.xlsx'.format(today)
    # # create data frame
    # df = pd.read_excel(x)
    #
    # # # selecting only important data for further investigation
    # cols = [x for x in range(0, 25)]
    #
    # # create new dataframe with only relevant columns
    # new_df = df.iloc[:, cols]
    #
    # # creating road dataframe
    # road = new_df.iloc[:, 0:25]
    #
    # # filter data
    # # road  = road.loc[:, road.columns != 'Bill?']
    # road = road.loc[:, road.columns != 'Leg 1, Billable']
    #
    # road = road.loc[:, road.columns != 'Leg 1, Carrier Code']
    #
    # road = road.loc[:, road.columns != 'Shipment']
    # # select only shipments that are without rates
    #
    # road = road.loc[road['Leg 1, Total costs'] == 0]
    # # select shipments that are unsolved
    # road = road.loc[road['Action taken'] != 'solved']
    #
    # # delete DSV DP Road cases
    # road = road.loc[road['Leg 1, Carrier Name'] != 'DSV DP Road']
    #
    # # delete shipments that are for Dmitrov
    # road = road.loc[road['Delivery city'] != 'Dmitrov']
    # # drop columns
    # road.drop(columns=['Action taken'], inplace=True)
    # # rename columns
    #
    # road.rename(columns={"Leg 1, Carrier Name": 'Carrier Name', "Leg 1, Transport Mode": "Transport mode"},
    #             inplace=True)
    # # dropping my columns
    # road.drop(columns=['Mail sent to IMS/key-user', 'date of Email', 'Email subject', 'Delay', 'Comments'],
    #           inplace=True)
    #
    # # reset index
    # road.reset_index(drop=True, inplace=True)
    #
    # # create excel object in pandas with proper formatting
    # filename = r'C:\Users\310295192\Desktop\Work\Rates\Road\missing rates for F&D\reports\\Road missing rates {}.xlsx'.format(
    #     today)
    # writer = pd.ExcelWriter(filename, engine='xlsxwriter', date_format='yyyy-mm-dd', datetime_format='yyyy-mm-dd')
    #
    # # dump data to excel
    # road.to_excel(writer, sheet_name='Missing_Rates', index=False)
    #
    # # prepare worksheet for working and the formating
    # workbook = writer.book
    #
    # # default cell format to size 10
    # workbook.formats[0].set_font_size(9)
    #
    # worksheet = writer.sheets['Missing_Rates']
    #
    # # set width of the column
    # worksheet.set_column('A:V', 16)
    #
    # # save file
    # writer.save()

    # from selenium import webdriver
    # import time
    # browser = webdriver.Chrome()
    # #getting to sharepoint
    # browser.get('https://share.philips.com/sites/STS020180301160509/Shared%20Documents/Forms/AllItems.aspx?id=%2Fsites%2FSTS020180301160509%2FShared%20Documents%2FPCT%20Poland%2FMJ')
    # # logging into webpage
    # browser.find_element_by_name('loginfmt').send_keys('maciej.janowski@philips.com')
    # from selenium.webdriver.common.keys import Keys
    # browser.find_element_by_name('loginfmt').send_keys(Keys.ENTER)
    # time.sleep(6)
    # # finding upload button
    # button = browser.find_element_by_xpath('//*[@id="appRoot"]/div/div[3]/div/div[2]/div[2]/div/div[3]/div/div/div/div/div/div/div/div/div[1]/div[2]/button/div/i[2]')
    # button.click()
    # time.sleep(2)
    # # finding hidden option for upload
    # button2 = browser.find_element_by_xpath("//input[@type='file']")
    # # uploading file to sharepoint
    # button2.send_keys(filename)

today = datetime.date.today()
# an overview for GTEC :)
NAM = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Road\US\reports\report road NAM {} updated.xlsx'.format(today))
FCL = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Sea\reports\FCL report {} updated.xlsx'.format(today))
LCL = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\LCL\LCL reports\DB LCL {} updated.xlsx'.format(today))
ROAD = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Road\reports complex\report road {} updated.xlsx'.format(today))
AIR = pd.read_excel(r'C:\Users\310295192\Desktop\Work\Rates\Air\reports\report AIR {} updated.xlsx'.format(today))

# adjusting the scope of columns
NAM = NAM[['Infodis','Action taken','date of Email','Email subject','Comments']]
FCL = FCL[['Infodis','Action taken','Pending/Solved','Date of email','Email subject','Comments']]
AIR = AIR[['Infodis','Action taken','Pending/Solved','date of email','E-mail subject','Comments']]
LCL = LCL[['Infodis','Action taken','Pending/Solved','Date of email','Email subject','Comments']]
ROAD = ROAD[['Infodis','Action taken','date of Email','Email subject','Comments']]

#adjusting the names of the columns

NAM.rename(columns={"date of Email":"Date of action","Action taken":"Pending/Solved"},inplace=True)
FCL.rename(columns={"Date of email":"Date of action"},inplace=True)
AIR.rename(columns={"date of email":"Date of action","E-mail subject":"Email subject"},inplace=True)
LCL.rename(columns={"Date of email":"Date of action"},inplace=True)
ROAD.rename(columns={"date of Email":"Date of action","Action taken":"Pending/Solved"},inplace=True)


# adjusting data from ROAD reports
ROAD.insert(loc=1, column='Action taken', value=np.select([ROAD['Pending/Solved'] =='solved',
                                                           ROAD['Pending/Solved'] =='pending'],['yes','yes'],default=np.nan))

NAM.insert(loc=1, column='Action taken', value=np.select([NAM['Pending/Solved'] =='solved',
                                                           NAM['Pending/Solved'] =='pending'],['yes','yes'],default=np.nan))
# concatanating DataFrames
TOTAL = pd.concat([NAM,FCL,AIR,LCL,ROAD])


# filtering data only pending and solved
TOTAL = TOTAL[(TOTAL["Pending/Solved"]=='solved')|(TOTAL['Pending/Solved']=='pending')]

# saving to excel file
ff = r'C:\Users\310295192\Desktop\Work\Rates\GTEC\shipments overview {} updated.xlsx'.format(today)
#TOTAL.to_excel(r'C:\Users\310295192\Desktop\SHIT HAPPENING.xlsx',index=False)
writer = pd.ExcelWriter(ff, engine='xlsxwriter',date_format = 'yyyy-mm-dd', datetime_format='yyyy-mm-dd')
#dump data to excel
TOTAL.to_excel(writer, sheet_name='Missing_Rates',index=False)
# save file
writer.save()
writer.close()
#TOTAL.to_excel(r'C:\Users\310295192\Desktop\Work\Rates\GTEC\shipments overview {} updated.xlsx'.format(today))
print(datetime.datetime.now())