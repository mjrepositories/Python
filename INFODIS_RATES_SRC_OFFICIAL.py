import pandas as pd
import numpy as np
from openpyxl import load_workbook
import shutil
from datetime import date
import string

# creating a list of letters for columns in files
alpha = list(string.ascii_lowercase)

# initial set-up for files
fcl_file = input('Provide path to FCL file. If not processed - enter "n"')
lcl_file = input("Provide path to LCL file. If not processed - enter 'n'")
air_file = input('Provide path to AIR file. If not processed enter "n"')
air_carrier = input("Provide letter for the carrier. a-  Expeditors,b - Nippon,"
                    "c - UPS AIRFREIGHT, d - CEVA or e - DB Schenker")

# selecting the name of the carrier
if air_carrier == 'a':
    carrier = 'Expeditors'
elif air_carrier == 'b':
    carrier = 'Nippon'
elif  air_carrier == 'c':
    carrier = 'UPS AIRFREIGHT'
elif air_carrier == 'd':
    carrier = 'CEVA'
elif air_carrier == 'e':
    carrier = 'DBS AIR'

if len(fcl_file) > 5:

    # reading file with FCL rates
    fcl_leg1 = pd.read_excel(fcl_file,sheet_name='LEG1')
    fcl_leg2 = pd.read_excel(fcl_file,sheet_name='LEG2')
    fcl_leg3 = pd.read_excel(fcl_file,sheet_name='LEG3')

    # creating list with SRC carriers
    src_carriers = ["ANL","APL","CMA","CNC","COSCO","HamburgSud","HAPAG","HYUNDAI","ICL","Maersk","ONE","ZIM","CH ROBINSON"]
    #  filtering data based on carriers list
    fcl_leg1 = fcl_leg1[fcl_leg1.CarrierName.isin(src_carriers)]
    fcl_leg2 = fcl_leg2[fcl_leg2.CarrierName.isin(src_carriers)]
    fcl_leg3 = fcl_leg3[fcl_leg3.CarrierName.isin(src_carriers)]


    # change code for CH Robinson
    d = {'CHRB5': 'CHRB6'}
    fcl_leg3 = fcl_leg3.replace(d)

    # starting and ending date for fcl rates
    datefrom = '2020-01-01'
    dateto ='2020-12-31'

    # inserting Zones columns
    fcl_leg1.insert(loc =5,column='FromZone',value = np.nan)
    fcl_leg1.insert(loc =9,column='ToZone',value = np.nan)

    # adjusting date format
    fcl_leg1['DateEffectiveFrom'] = fcl_leg1['DateEffectiveFrom'].dt.date
    fcl_leg1['DateEffectiveTo'] = fcl_leg1['DateEffectiveTo'].dt.date

    # inserting Zones columns
    fcl_leg2.insert(loc =5,column='FromZone',value = np.nan)
    fcl_leg2.insert(loc =9,column='ToZone',value = np.nan)

    # adjusting date format
    fcl_leg2['DateEffectiveFrom'] = fcl_leg2['DateEffectiveFrom'].dt.date
    fcl_leg2['DateEffectiveTo'] = fcl_leg2['DateEffectiveTo'].dt.date

    # inserting Zones columns
    fcl_leg3.insert(loc =5,column='FromZone',value = np.nan)
    fcl_leg3.insert(loc =9,column='ToZone',value = np.nan)

    # adjusting date format
    fcl_leg3['DateEffectiveFrom'] = fcl_leg3['DateEffectiveFrom'].dt.date
    fcl_leg3['DateEffectiveTo'] = fcl_leg3['DateEffectiveTo'].dt.date

    # checking the today date
    today = date.today().strftime('%Y-%m-%d')
    for cr in src_carriers:
        # creating name of the template for adding sheets with rates for SRC
        src_fcl_dir = r'C:\Users\310295192\Desktop\Python\Projects\RATES_FOR_SRC\FCL_template_SRC.xlsx'
        dst_fcl_dir = r'C:\Users\310295192\Desktop\Work\Rates\SRC\Preparing FCL rates\2010 FCL rates {} SRC template {}.xlsx'.format(today,cr)

        # coping template for inserting adjusted rates
        shutil.copy(src_fcl_dir,dst_fcl_dir)

        path = dst_fcl_dir
        book = load_workbook(path)

        # assigning workbook to variable
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book


        # filtering data for specific carrier

        fcl_leg1_cr = fcl_leg1[fcl_leg1.CarrierName == cr]
        fcl_leg2_cr = fcl_leg2[fcl_leg2.CarrierName == cr]
        fcl_leg3_cr = fcl_leg3[fcl_leg3.CarrierName == cr]
        # adding each leg to the workbook
        fcl_leg1_cr.to_excel(writer, sheet_name = 'PRE_leg1',index=False)
        fcl_leg2_cr.to_excel(writer, sheet_name = 'FCL_leg2',index=False)
        fcl_leg3_cr.to_excel(writer, sheet_name = 'ONC_leg3',index=False)

        worksheet1 = writer.sheets['PRE_leg1']
        worksheet2 = writer.sheets['FCL_leg2']
        worksheet3 = writer.sheets['ONC_leg3']
        # for x in alpha:
        #     # setting up the width of the columns
        #     worksheet1.column_dimensions[x.upper()].width = 20
        #
        #     worksheet2.column_dimensions[x.upper()].width = 20
        #
        #     worksheet3.column_dimensions[x.upper()].width = 20
        #
        # saving and closing
        writer.save()
        writer.close()


if len(lcl_file) > 5:
    # reading lcl file
    lcl_leg1 = pd.read_excel(lcl_file, sheet_name='Pre_leg1')
    lcl_leg2 = pd.read_excel(lcl_file,sheet_name='LCL_leg2')
    lcl_leg3 = pd.read_excel(lcl_file,sheet_name='ONC_leg3')

    # inserting Zones columns, service level and carrier name
    lcl_leg1.insert(loc=5, column='FromZone', value=np.nan)
    lcl_leg1.insert(loc=9, column='ToZone', value=np.nan)
    lcl_leg1.insert(loc=13, column='ServiceLevel', value=np.nan)
    lcl_leg1.CarrierName = 'DBS LCL'

    # adjusting date format
    lcl_leg1['DateEffectiveFrom'] = lcl_leg1['DateEffectiveFrom'].dt.date
    lcl_leg1['DateEffectiveTo'] = lcl_leg1['DateEffectiveTo'].dt.date

    # inserting Zones columns, service level and carrier name
    lcl_leg2.insert(loc=5, column='FromZone', value=np.nan)
    lcl_leg2.insert(loc=9, column='ToZone', value=np.nan)
    lcl_leg2.insert(loc=13, column='ServiceLevel', value=np.nan)
    lcl_leg2.CarrierName = 'DBS LCL'

    # adjusting date format
    lcl_leg2['DateEffectiveFrom'] = lcl_leg2['DateEffectiveFrom'].dt.date
    lcl_leg2['DateEffectiveTo'] = lcl_leg2['DateEffectiveTo'].dt.date

    # inserting Zones columns, service level and carrier name
    lcl_leg3.insert(loc=5, column='FromZone', value=np.nan)
    lcl_leg3.insert(loc=9, column='ToZone', value=np.nan)
    lcl_leg3.insert(loc=13, column='ServiceLevel', value=np.nan)
    lcl_leg3.CarrierName = 'DBS LCL'

    # adjusting date format
    lcl_leg3['DateEffectiveFrom'] = lcl_leg3['DateEffectiveFrom'].dt.date
    lcl_leg3['DateEffectiveTo'] = lcl_leg3['DateEffectiveTo'].dt.date

    today = date.today().strftime('%Y-%m-%d')
    # creating name of the template for adding sheets with rates for SRC
    src_lcl_dir = r'C:\Users\310295192\Desktop\Python\Projects\RATES_FOR_SRC\DB_template_SRC.xlsx'
    dst_lcl_dir = r'C:\Users\310295192\Desktop\Work\Rates\SRC\Uploads\2019 DB Schenker LCL rates {} SRC template.xlsx'.format(today)

    # coping template for inserting adjusted rates
    shutil.copy(src_lcl_dir, dst_lcl_dir)

    path = dst_lcl_dir
    book = load_workbook(path)
    # assinging workbook to variable
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book

    # adding each leg to the workbook
    lcl_leg1.to_excel(writer, sheet_name='PRE_leg1', index=False)
    lcl_leg2.to_excel(writer, sheet_name='LCL_leg2', index=False)
    lcl_leg3.to_excel(writer, sheet_name='ONC_leg3', index=False)

    worksheet1 = writer.sheets['PRE_leg1']
    worksheet2 = writer.sheets['LCL_leg2']
    worksheet3 = writer.sheets['ONC_leg3']
    for x in alpha:
        # setting up the width of the columns
        worksheet1.column_dimensions[x.upper()].width = 20

        worksheet2.column_dimensions[x.upper()].width = 20

        worksheet3.column_dimensions[x.upper()].width = 20

    # saving and closing
    writer.save()
    writer.close()

if len(air_file) > 5 and air_carrier in ['a','b','c','d']:
    # preparing rates file for AIR
    # reading lcl file
    air_leg1 = pd.read_excel(air_file, sheet_name='PRE_leg1')
    air_leg2 = pd.read_excel(air_file, sheet_name='AIR_leg2')
    air_leg3 = pd.read_excel(air_file, sheet_name='ONC_leg3')
    # inserting Zones columns
    air_leg1.insert(loc=5, column='FromZone', value=np.nan)
    air_leg1.insert(loc=9, column='ToZone', value=np.nan)

    # adjusting date format
    air_leg1['DateEffectiveFrom'] = air_leg1['DateEffectiveFrom'].dt.date
    air_leg1['DateEffectiveTo'] = air_leg1['DateEffectiveTo'].dt.date

    # inserting Zones columns
    air_leg2.insert(loc=5, column='FromZone', value=np.nan)
    air_leg2.insert(loc=9, column='ToZone', value=np.nan)

    # adjusting date format
    air_leg2['DateEffectiveFrom'] = air_leg2['DateEffectiveFrom'].dt.date
    air_leg2['DateEffectiveTo'] = air_leg2['DateEffectiveTo'].dt.date

    # inserting Zones columns
    air_leg3.insert(loc=5, column='FromZone', value=np.nan)
    air_leg3.insert(loc=9, column='ToZone', value=np.nan)

    # adjusting date format
    air_leg3['DateEffectiveFrom'] = air_leg3['DateEffectiveFrom'].dt.date
    air_leg3['DateEffectiveTo'] = air_leg3['DateEffectiveTo'].dt.date

    # selecting the name of the carrier
    if air_carrier == 'b':
        air_leg1['CarrierName'] = 'Nippon'
        air_leg2['CarrierName'] = 'Nippon'
        air_leg3['CarrierName'] = 'Nippon'
    elif air_carrier == 'c':
        carrier = 'UPS AIRFREIGHT'
        air_leg1['CarrierName'] = 'UPS AIRFREIGHT'
        air_leg2['CarrierName'] = 'UPS AIRFREIGHT'
        air_leg3['CarrierName'] = 'UPS AIRFREIGHT'
    elif air_carrier == 'd':
        carrier = 'CEVA'
        air_leg1['CarrierName'] = 'CEVA'
        air_leg2['CarrierName'] = 'CEVA'
        air_leg3['CarrierName'] = 'CEVA'
    elif air_carrier == 'd':
        carrier = 'DBS AIR'
        air_leg1['CarrierName'] = 'DBS AIR'
        air_leg2['CarrierName'] = 'DBS AIR'
        air_leg3['CarrierName'] = 'DBS AIR'

    # avoinding duplicates
    air_leg1['duplicate'] = air_leg1['FromUNLOCODE'] + "_" + air_leg1['ToUNLOCODE'] + '_' + air_leg1[
        'ServiceLevel'].astype(str)
    air_leg2['duplicate'] = air_leg2['FromUNLOCODE'] + "_" + air_leg2['ToUNLOCODE'] + '_' + air_leg2[
        'ServiceLevel'].astype(str) + '_' + air_leg2['ChargeType']
    air_leg3['duplicate'] = air_leg3['FromUNLOCODE'] + "_" + air_leg3['ToUNLOCODE'] + '_' + air_leg3[
        'ServiceLevel'].astype(str)

    # removing duplicates from each leg
    air_leg1.drop_duplicates(subset='duplicate', inplace=True)

    air_leg2.drop_duplicates(subset='duplicate', inplace=True)

    air_leg3.drop_duplicates(subset='duplicate', inplace=True)

    # dropping duplicates columns
    air_leg1.drop(columns=['duplicate'], inplace=True)
    air_leg2.drop(columns=['duplicate'], inplace=True)
    air_leg3.drop(columns=['duplicate'], inplace=True)

    today = date.today().strftime('%Y-%m-%d')
    # creating name of the template for adding sheets with rates for SRC
    src_air_dir = r'C:\Users\310295192\Desktop\Python\Projects\RATES_FOR_SRC\AIR_template_SRC.xlsx'
    dst_air_dir = r'C:\Users\310295192\Desktop\Work\Rates\SRC\Uploads\2019-2020 {} rates {} SRC template.xlsx'.format(
        air_carrier,today)

    # coping template for inserting adjusted rates
    shutil.copy(src_air_dir, dst_air_dir)

    path = dst_air_dir
    book = load_workbook(path)
    # assinging workbook to variable
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book

    # adding each leg to the workbook
    air_leg1.to_excel(writer, sheet_name='PRE_leg1', index=False)
    air_leg2.to_excel(writer, sheet_name='AIR_leg2', index=False)
    air_leg3.to_excel(writer, sheet_name='ONC_leg3', index=False)

    worksheet1 = writer.sheets['PRE_leg1']
    worksheet2 = writer.sheets['AIR_leg2']
    worksheet3 = writer.sheets['ONC_leg3']
    for x in alpha:
        # setting up the width of the columns
        worksheet1.column_dimensions[x.upper()].width = 20

        worksheet2.column_dimensions[x.upper()].width = 20

        worksheet3.column_dimensions[x.upper()].width = 20

    # saving and closing
    writer.save()
    writer.close()