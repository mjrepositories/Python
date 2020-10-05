from selenium import webdriver
import time
import openpyxl
car_name = input ("Provide carrier name: ").upper()
filenaming = r'C:\Users\310295192\Desktop\Python\Projects\Rates\AIR\{}\{}.xlsx'.format(car_name,car_name)
# setting up list for mjb numbers
wb =openpyxl.load_workbook(filenaming)
#assigns variable to worksheet
sheet = wb['Costs']
#finds last row
maximum=sheet.max_row
#setting an empty set
mjb_list = []
#going through the list on all jobs
for xls_row in range(2,maximum+1):
    mjb_no = sheet.cell(row=xls_row,column=1).value
    print(mjb_no)
    # checking if rate are already verified
    checking = sheet.cell(row=xls_row,column=32).value
    if checking !='yes' and mjb_no != None:
        mjb_list.append(mjb_no)


print(mjb_list)
print(len(mjb_list))

wb.close()
browser = webdriver.Chrome()
# getting to Optilo
browser.get("https://ot3.optilo.eu/opt_ext_smxc9a/p001/main.php?m=cwlib&c=login&return_url=main.php%3Fm%3Dcwlib%26c%3Dstartpage")
browser.maximize_window()
# loggin in
emailElem = browser.find_element_by_id('inpLogin')
emailElem.send_keys('maciej.janowski@philips.com')
# inputting password
password = browser.find_element_by_id('inpPassword')
password.send_keys('Maciej0312@')

# submitting logon details
buttonlog = browser.find_element_by_id('submitLogin')
buttonlog.click()

for every in mjb_list:
    try:
        # finding multi tab
        # multi = browser.find_element_by_id('menu-258664')
        multi = browser.find_element_by_id('menu-259381')
        multi.click()
        # finding multi job tab
        # joblist = browser.find_element_by_id('menu-258666')
        joblist = browser.find_element_by_id('menu-259383')
        joblist.click()
        # finding box for job number
        jobnumber = browser.find_element_by_id('DF650000161_number')
        jobnumber.clear()
        # looping through each job number
        jobnumber.send_keys(every)
        time.sleep(3)
        # finding search button
        # searching = browser.find_element_by_id('action[650000161][1529675]')
        searching = browser.find_element_by_id('action[650000161][3242]')
        searching.click()
        # finding button button and clicking it
        time.sleep(2)
        buttonforjob = browser.find_element_by_link_text(every)
        buttonforjob.click()
        time.sleep(3)
        # finding transport service type field
        type = browser.find_element_by_id('DF650000229_transport_service_type_box')
        # getting to possible options
        options_sl = type.find_elements_by_tag_name("option")
        # looping through options
        lets_count = 1
        counting =1
        for selection in options_sl:
            selected = selection.get_attribute('selected')
            if selected == None:
                lets_count += 1
            else:
                counting = lets_count
                print(lets_count)


        if counting == 2:
            service_type = 'D2D'
        elif counting == 3:
            service_type = 'D2P'
        else:
            service_type = 'D2D'
        print(service_type)
        # getting to parameters & measures
        job_table = browser.find_element_by_id('ZrodloKontekst-Job')
        # finding all options in job table
        table_content = job_table.find_elements_by_tag_name('li')
        for job_option in table_content:
            # if "parameters & measures are found - it is clicking it
            if job_option.text == 'Parameters & Measures':
                job_option.click()
        # switch focus to current window
        new_window = browser.find_element_by_id('TB_iframeContent')
        browser.switch_to.frame(new_window)

        print('i am here')
        # finding parameters table
        par_tab = browser.find_element_by_id("ZrodloDane-Parametr")
        # getting directly to table with each row
        #par_tab_rows = par_tab.find_element_by_xpath('//*[@id="DG520000112"]')
        # finding the number of rows in table
        row_c = len(par_tab.find_elements_by_xpath('//*[@id="DG520000112"]/tbody/tr'))
        row = 1
        for x in range(2,row_c):
            checking = par_tab.find_element_by_xpath('//*[@id="DG520000112"]/tbody/tr[{}]/td[2]'.format(x))
            # checking the input tags
            value_added = checking.find_elements_by_tag_name('input')
            for element in value_added:
                # checking the value attribute for every input tag
                valuable = element.get_attribute('value')
                # if value contains SLA then we have the row for service level
                if "Carrier Service Level" in valuable:
                    row = x
        print(row)
        # looking for proper service level based on extract row number
        # looking for field with service level
        checking = browser.find_element_by_xpath('//*[@id="DG520000112"]/tbody/tr[{}]/td[3]'.format(row))
        # checking every option tag
        options_service_level = checking.find_elements_by_tag_name('option')
        counter = 1
        rowing = 1
        for x in options_service_level:
            # checking for selected option with option tag
            service = x.get_attribute('selected')
            # if the selection is found - it returns the number of row
            if service == None:
                counter += 1
            else:
                rowing = counter


        seek_value = browser.find_element_by_xpath('//*[@id="DG520000112"]/tbody/tr[{}]/td[3]/select/option[{}]'.format(row,rowing))
        print('next step')
        service_level = ""
        # look for the appropriate service level and assigns it to variable
        if "SL1" in seek_value.text:
            service_level = "SL1"
        elif "SL2" in seek_value.text:
            service_level = "SL2"
        elif "SL3" in seek_value.text:
            service_level = "SL3"
        elif "SL4" in seek_value.text:
            service_level = "SL4"
        elif "SL5" in seek_value.text:
            service_level = "SL5"
        elif "SL6" in seek_value.text:
            service_level = "SL6"
        elif "SL7" in seek_value.text:
            service_level = "SL7"
        print(service_level)
        # going again to side (i was not able to close this f**** window)
        current_address = browser.current_url
        browser.get(current_address)

        # getting port of loading
        port_start = browser.find_element_by_id('DF650000163_start_id_krt_port')
        port_start_val = port_start.get_attribute('value')
        air_start = port_start_val

        # getting port of discharge
        port_end = browser.find_element_by_id('DF650000163_end_id_krt_port')
        port_end_val = port_end.get_attribute('value')
        air_stop = port_end_val
        # taking the current address info
        current_address = browser.current_url
        # going to multi packages tab
        web_address = str(current_address)
        # replace jobdetails to jobcosts and go the site with costs
        newaddress = web_address.replace('JobDetails', 'JobPackages')
        # going to new files tab
        browser.get(newaddress)
        time.sleep(3)

        # checking the volume
        vol = browser.find_element_by_xpath('//*[@id="DL650000369"]/tbody/tr[2]/td[2]').text
        volume = float(vol.replace(",", "."))

        chargeable = round(volume * 166.67,2)

        # checking the weight
        wht = browser.find_element_by_xpath('//*[@id="DL650000369"]/tbody/tr[2]/td[3]').text
        weight = round(float(wht.replace(",", ".")),2)

        if chargeable > weight:
            right_weight = chargeable
        elif chargeable < weight:
            right_weight = weight

        print(service_type)
        print(service_level)
        print(air_start)
        print(air_stop)
        print(volume)
        print(chargeable)
        print(weight)
        print(right_weight)

        # opening workbook
        wb = openpyxl.load_workbook(filenaming)
        # assigns variable to worksheet
        sheet = wb['Costs']
        # finds last row
        maximum = sheet.max_row
        for xls_row in range(2, maximum + 1):
            mjb_no = sheet.cell(row=xls_row, column=1).value
            print(mjb_no)
            if mjb_no == every:
            # checking if rate are already verified
                verification = sheet.cell(row=xls_row, column=32).value
                if verification != 'yes':
                    print('verifcation done')
                    # it assigns vales that program found out in Optilo
                    # and assigns shipment as verified
                    sheet.cell(row=xls_row, column=2).value = service_level
                    print('first value assigned')
                    sheet.cell(row=xls_row, column=3).value = service_type
                    print('second value assigned')
                    sheet.cell(row=xls_row, column=4).value = air_start
                    print('third value assigned')
                    sheet.cell(row=xls_row, column=5).value = air_stop
                    print('fourth value assigned')
                    sheet.cell(row=xls_row, column=9).value = volume
                    print('fifth value assigned')
                    sheet.cell(row=xls_row, column=10).value = weight
                    print('sixth value assigned')
                    sheet.cell(row=xls_row, column=11).value = right_weight
                    print('seventh value assigned')
                    sheet.cell(row=xls_row, column=32).value = 'yes'
                    print('eighth value assigned')
                    break
        wb.save(filenaming)
        wb.close()
    # https://ot3.optilo.eu/opt_ext_smxc9a/p001/druid.php?m=multi&s=MultiJobDetails&key=id_mlt_job,17429
    except:
        print('something went wrong')

browser.close()





