from selenium import webdriver
import time
import os
import openpyxl



wb= openpyxl.load_workbook(r'C:\Users\310295192\Desktop\Python\Projects\REPROCESSING_COSTS_OPTILO\reprocess.xlsx')
#assigns variable to worksheet
sheet = wb['sheet1']
#finds last row
maximum=sheet.max_row
#setting an empty set
jobs = set()
#going through the list on all jobs
for x in range(2,maximum+1):
    jobs.add(sheet.cell(row=x,column=1).value)

print(jobs)
wb.close()

#looping through the every value
browser = webdriver.Chrome()
#loggin into account
browser.maximize_window()
browser.get("https://ot3.optilo.eu/opt_ext_smxc9a/p001/main.php?m=cwlib&c=login&return_url=main.php%3Fm%3Dcwlib%26c%3Dstartpage")
emailElem = browser.find_element_by_id('inpLogin')
emailElem.send_keys('maciej.janowski@philips.com')
passwordElem = browser.find_element_by_id('inpPassword')
passwordElem.send_keys('Maciej0312@')
login = browser.find_element_by_id('submitLogin')
login.click()
for x in jobs:
    # opening multi list
    multi = browser.find_element_by_id('menu-259381')
    multi.click()
    # picking up jobs list
    joboverview = browser.find_element_by_id('menu-259383')
    joboverview.click()
    time.sleep(5)

    # ,'MJB9019469','MJB9019468','MJB9019464

    # finding box for job number
    jobnumber = browser.find_element_by_id('DF650000161_number')
    jobnumber.clear()
    # looping through each job number
    jobnumber.send_keys(x)
    time.sleep(3)
    # finding search button
    searching = browser.find_element_by_id('action[650000161][3242]')
    searching.click()
    # finding button button and clicking it
    time.sleep(2)
    buttonforjob = browser.find_element_by_link_text(x)
    buttonforjob.click()
    time.sleep(3)
    # copy address from tab
    current_address = browser.current_url
    # turn to string
    web_address = str(current_address)
    # replace jobdetails to jobcosts and go the site with costs
    newaddress = web_address.replace('JobDetails', 'JobCosts')
    browser.get(newaddress)
    time.sleep(5)
    # finding the recalculation button and clicks it
    recalculate = browser.find_element_by_name('action[650000281][895]')
    recalculate.click()
    time.sleep(3)
browser.close()




