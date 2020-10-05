import pandas as pd
import openpyxl

dating = input('Provide date: ')
# reading data
contracts = pd.read_excel(r"C:\Users\310295192\Desktop\Python\Projects\MJ_LNOS\contracts.xlsx",sheet_name='rates',dtype={'zipto':str})
diesel = pd.read_excel(r"C:\Users\310295192\Desktop\Python\Projects\MJ_LNOS\contracts.xlsx",sheet_name='diesel')
fuel = pd.read_excel(r"C:\Users\310295192\Desktop\Python\Projects\MJ_LNOS\contracts.xlsx",sheet_name='fuel')

# setting up diesel price
diesel_price = diesel.iloc[0,0]

# setting filter for fuel
diesel_up = diesel_price +0.02
diesel_down =  diesel_price -0.02

# getting fuel table with limited rows
fuel = fuel[(fuel.NAFP>diesel_down)&(fuel.NAFP<diesel_up)]

# rouding fuel
fuel.NAFP = fuel.round({'NAFP':3})

# getting fuel price per mile
ppm = 0
for index,row in fuel.iterrows():
    if (diesel_price>=row['NAFP']) ==False:
        ppm = round(row['Rounded FSC'],2)
        break

# calculating surcharge
contracts['surcharge'] = contracts['miles'] *ppm

# creating id for lanes
contracts['id'] = contracts['carrier'] + contracts['zipfrom'].astype(str)

# creating list for looping
car_list =list(set(contracts.id.to_list()))

# looping through each lane id
for car in car_list:
    filename = r"C:\Users\310295192\Desktop\Python\Projects\MJ_LNOS\template.xlsx"
    frame = contracts[contracts.id == car]
    x = 0
    # open worksheet
    wb = openpyxl.load_workbook(filename)
    sheet = wb['Freight']
    sheet.cell(row=3, column=3).value = frame.iloc[0, 2] + " " + str(frame.iloc[0, 3])
    sheet.cell(row=5, column=3).value = frame.iloc[0, 5]
    sheet.cell(row=6, column=3).value = frame.iloc[0, 1]
    sheet.cell(row=8, column=3).value = dating

    for index, row in frame.iterrows():
        # FREIGHT CANCULATION

        # sheet = wb['Freight']
        # sheet.cell(row=24, column=4 + x).value = frame.iloc[0 + x, 7]
        # sheet.cell(row=28, column=4 + x).value = frame.iloc[0 + x, 4]
        # sheet.cell(row=32, column=4 + x).value = frame.iloc[0 + x, 8]
        #
        # # FUEL CALCULATION
        # sheet = wb['Surcharge']
        #
        # sheet.cell(row=36, column=4 + x).value = frame.iloc[0 + x, 4]
        # sheet.cell(row=40, column=4 + x).value = frame.iloc[0 + x, 9]
        # x += 1
        sheet = wb['Freight']
        sheet.cell(row=24, column=4 + x).value = row['leadtime']
        sheet.cell(row=28, column=4 + x).value = row['zipto']
        sheet.cell(row=32, column=4 + x).value = row['charge']

        # FUEL CALCULATION
        sheet = wb['Surcharge']

        sheet.cell(row=36, column=4 + x).value = row['zipto']
        sheet.cell(row=40, column=4 + x).value = row['surcharge']
        x += 1



    filename = r"C:\Users\310295192\Desktop\testingroadUS {}.xlsx".format(car)
    wb.save(filename)
    wb.close()
