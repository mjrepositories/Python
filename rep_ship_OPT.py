from selenium import webdriver
import time
import os

browser=webdriver.Chrome()
#getting to Optilo site
browser.get("https://ot3.optilo.eu/opt_ext_smxc9a/p001/main.php?m=cwlib&c=login&return_url=main.php%3Fm%3Dcwlib%26c%3Dstartpage")
browser.maximize_window()
#looking for login field and typing in the username
log = browser.find_element_by_id('inpLogin')
log.send_keys('maciej.janowski@philips.com')
#looking for login field and typing in the password to accoutn
word = browser.find_element_by_id('inpPassword')
word.send_keys('Maciej0312@')
#looking for submit button to get to the system and confirming input details
login = browser.find_element_by_id('submitLogin')
login.click()
#looking for finances tab and unwailing it
finances = browser.find_element_by_id('menu-258134')
finances.click()
#looking for invoice report and opening it
invoice = browser.find_element_by_id('menu-258139')
invoice.click()

time.sleep(13)
#looking for pick-up date from and filling in the values
pickup=browser.find_element_by_id("DF100007350_pick_up_date_from")
pickup.send_keys('2019-01-10')
time.sleep(1)
#looking for unloading date to and filling in the values
unloading = browser.find_element_by_id('DF100007350_delivery_date_until')
unloading.send_keys('2019-01-18')
time.sleep(1)
#finding search button and searching the results
searching = browser.find_element_by_id('action[100007350][1526400]')
searching.click()
time.sleep(11)
#finding XLS button and clicking it
excel = browser.find_element_by_class_name('fieldset-links')
# backup button =excel.find_element_by_xpath("//a[@href='#']")
#backup 2 button =excel.find_element_by_xpath("//a[@onclick='druid.fileDownload('main.php?m=cwlib&amp']")
#button =excel.find_element_by_xpath("//a[@href='#']")
#clicking the href by finding XLS.... FUCK YEAH!!!
button = browser.find_element_by_link_text("XLS")
button.click()
time.sleep(10)
#opening the report with jobs
import pyautogui
#Point(x=128, y=1015) - downloaded file
#Point(x=531, y=50) - dialog window
#Point(x=790, y=510) - save button
#Opens the excel file after download
pyautogui.click(128,1015)
time.sleep(6)
browser.close()
#closes the browser
time.sleep(6)
#click f12 for "save as"
pyautogui.hotkey('f12')
time.sleep(2)
#type in the name of the file
pyautogui.typewrite('reprocess')
time.sleep(2)
#click the address window
pyautogui.click(530,50)
time.sleep(2)
#type in the folder where it should be saved and click applies it
pyautogui.typewrite(r'C:\Users\310295192\Desktop\VBA projects\Key-users app\reprocessing')
pyautogui.typewrite(["enter"])
#Point(x=212, y=428) for picking up the menu
time.sleep(2)
pyautogui.click(212,428)
#Point(x=256, y=445) for picking up the file extension
time.sleep(2)
pyautogui.click(256,445)
time.sleep(2)
#saves file under new name in specified folder
pyautogui.click(182,400)
pyautogui.typewrite(["enter"])
time.sleep(2)
#closes the window
pyautogui.click(1896,13)
#remove duplicates
import openpyxl

wb= openpyxl.load_workbook(r'C:\Users\310295192\Desktop\VBA projects\Key-users app\reprocessing\reprocess.xlsx')
#assigns variable to worksheet
sheet = wb['MltJobCostReport']
#finds last row
maximum=sheet.max_row
#setting an empty set
jobs = set()
#going through the list on all jobs
for x in range(2,maximum+1):
    jobs.add(sheet.cell(row=x,column=1).value)
#looping through the every value
browser = webdriver.Chrome()
#loggin into account
browser.maximize_window()
browser.get("https://ot3.optilo.eu/opt_ext_smxc9a/p001/main.php?m=cwlib&c=login&return_url=main.php%3Fm%3Dcwlib%26c%3Dstartpage")
emailElem = browser.find_element_by_id('inpLogin')
emailElem.send_keys('maciej.janowski@philips.com')
passwordElem = browser.find_element_by_id('inpPassword')
passwordElem.send_keys('Maciej0312@')
login = browser.find_element_by_id('submitLogin')
login.click()
for x in jobs:
    # opening multi list
    multi = browser.find_element_by_id('menu-259381')
    multi.click()
    # picking up jobs list
    joboverview = browser.find_element_by_id('menu-259383')
    joboverview.click()
    time.sleep(5)

    # ,'MJB9019469','MJB9019468','MJB9019464

    # finding box for job number
    jobnumber = browser.find_element_by_id('DF650000161_number')
    jobnumber.clear()
    # looping through each job number
    jobnumber.send_keys(x)
    time.sleep(3)
    # finding search button
    searching = browser.find_element_by_id('action[650000161][3242]')
    searching.click()
    # finding button button and clicking it
    time.sleep(2)
    buttonforjob = browser.find_element_by_link_text(x)
    buttonforjob.click()
    time.sleep(3)
    # copy address from tab
    current_address = browser.current_url
    # turn to string
    web_address = str(current_address)
    # replace jobdetails to jobcosts and go the site with costs
    newaddress = web_address.replace('JobDetails', 'JobCosts')
    browser.get(newaddress)
    time.sleep(5)
    # finding the recalculation button and clicks it
    recalculate = browser.find_element_by_name('action[650000281][895]')
    recalculate.click()
    time.sleep(3)
browser.close()
os.remove(r'C:\Users\310295192\Desktop\VBA projects\Key-users app\reprocessing\reprocess.xlsx')



