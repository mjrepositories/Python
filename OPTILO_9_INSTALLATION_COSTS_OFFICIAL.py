from selenium import webdriver
import time
import openpyxl
car_name = input ("Provide carrier name: ").title()
filenaming = r'C:\Users\310295192\Desktop\Python\Projects\Rates\INSTALLATION\LMP\{}\{}.xlsx'.format(car_name,car_name)
# setting up list for mjb numbers
wb =openpyxl.load_workbook(filenaming)
#assigns variable to worksheet
sheet = wb['Costs']
#finds last row
maximum=sheet.max_row
#setting an empty set
mjb_list = []
#going through the list on all jobs
for xls_row in range(2,maximum+1):
    mjb_no = sheet.cell(row=xls_row,column=1).value
    print(mjb_no)
    # checking if rate are already verified
    checking = sheet.cell(row=xls_row,column=14).value
    if checking !='yes' and mjb_no != None:
        mjb_list.append(mjb_no)



print(mjb_list)
print(len(mjb_list))

wb.close()
browser = webdriver.Chrome()
# getting to Optilo
browser.get("https://ot3.optilo.eu/opt_ext_smxc9a/p001/main.php?m=cwlib&c=login&return_url=main.php%3Fm%3Dcwlib%26c%3Dstartpage")
browser.maximize_window()
# loggin in
emailElem = browser.find_element_by_id('inpLogin')
emailElem.send_keys('maciej.janowski@philips.com')
# inputting password
password = browser.find_element_by_id('inpPassword')
password.send_keys('Maciej0312@')

# submitting logon details
buttonlog = browser.find_element_by_id('submitLogin')
buttonlog.click()

for every in mjb_list:
    try:
        # finding multi tab
        # multi = browser.find_element_by_id('menu-258664')
        multi = browser.find_element_by_id('menu-259381')
        multi.click()
        # finding multi job tab
        # joblist = browser.find_element_by_id('menu-258666')
        joblist = browser.find_element_by_id('menu-259383')
        joblist.click()
        # finding box for job number
        jobnumber = browser.find_element_by_id('DF650000161_number')
        jobnumber.clear()
        # looping through each job number
        jobnumber.send_keys(every)
        time.sleep(3)
        # finding search button
        # searching = browser.find_element_by_id('action[650000161][1529675]')
        searching = browser.find_element_by_id('action[650000161][3242]')
        searching.click()
        # finding button button and clicking it
        time.sleep(2)
        buttonforjob = browser.find_element_by_link_text(every)
        buttonforjob.click()
        time.sleep(3)
        # check current address
        current_address = browser.current_url
        # going to lines tab
        web_address = str(current_address)
        # replace jobdetails to joblines and go the site with costs
        newaddress = web_address.replace('JobDetails', 'JobLines')
        # going to new files tab
        browser.get(newaddress)
        time.sleep(3)
        mjb_count_row = len(browser.find_elements_by_xpath("//table[@id='DL650000243']/tbody/tr"))
        time.sleep(2)
        print(mjb_count_row)
        # looping through each row for order lines
        for y in range(2, mjb_count_row + 1):
            # extracting the number of rows in table for order lines
            dq_tag = browser.find_element_by_xpath('//table[@id="DL650000243"]/tbody/tr[{}]/td[35]'.format(y)).text
            if dq_tag == "Y":
                dq_tag = "Yes"
                p_no = browser.find_element_by_xpath('//table[@id="DL650000243"]/tbody/tr[{}]/td[3]'.format(y)).text
                p_name = browser.find_element_by_xpath('//table[@id="DL650000243"]/tbody/tr[{}]/td[4]'.format(y)).text
                break
            else:
                dq_tag = "No"
        # opening workbook
        wb = openpyxl.load_workbook(filenaming)
        # assigns variable to worksheet
        sheet = wb['Costs']
        # finds last row
        maximum = sheet.max_row
        for xls_row in range(2, maximum + 1):
            mjb_no = sheet.cell(row=xls_row, column=1).value
            if mjb_no == every:
                # checking if rate are already verified
                verification = sheet.cell(row=xls_row, column=14).value
                if verification != 'yes':
                    print('verifcation done')
                    # it assigns vales that program found out in Optilo
                    # and assigns shipment as verified
                    sheet.cell(row=xls_row, column=2).value = dq_tag
                    print('first value assigned')
                    sheet.cell(row=xls_row, column=3).value = p_no
                    print('second value assigned')
                    sheet.cell(row=xls_row, column=4).value = p_name
                    print('third value assigned')
                    sheet.cell(row=xls_row, column=14).value = 'yes'
                    print('checked')
                    break
        wb.save(filenaming)
        wb.close()

    except:
        print("something went wrong")

browser.close()