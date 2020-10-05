import tkinter as tk
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from selenium import webdriver
import openpyxl
import PyPDF2
import time
wysok= 300
szer = 400

root = tk.Tk()

def open_infodis():
    # we are opening webrowser here
    global browser
    browser = webdriver.Chrome()

    browser.get("https://www.infodis.net/philips/net/LogOn/Authenticate?returnUrl=https%3a%2f%2fwww.infodis.net%2fphilips%2fnet%2f")
    for x in range(1,8):
        print(x)
#this is the function that we will use to add packages


def add_packages():
    ftypes = [('All files', '*.*')]
    ttl = "Select the file"
    dir1 = r'C:\Users\310295192\Desktop\Python\Projects'
    root.fileName = askopenfilename(filetypes=ftypes, initialdir=dir1, title=ttl)
    naming = str(root.fileName)
    print(naming)

    # here i need a procedure that will create excel fiel from pdf
    #and then extract the data from excel file to the fields in infodis

    # otwiera plik do zapisania
    file = open(naming, 'rb')
    # otwiera pdf
    pdfreader = PyPDF2.PdfFileReader(file)
    # dostaje sie do stron
    print(pdfreader.getNumPages())
    # otwiera pierwszą stronę
    paging = pdfreader.getPage(1)
    # wyciąga dane z pierwszej strony
    strona = paging.extractText()
    # zamienia kropki na przecinki
    nowastrona = strona.replace(',', '.')
    print(nowastrona)
    # znajduje początek listy
    start_pack = nowastrona.find('/001')
    print("Customer" in nowastrona)
    # znalazlo na 83
    # znajduje koniec listy
    end_pack = nowastrona.find("Customer")
    # tworzy zmienna tylko z danymi odnosnie paczek
    products = nowastrona[start_pack + 1:end_pack]
    print(start_pack)
    print(end_pack)
    print(products)
    wb = openpyxl.load_workbook(r"C:\Users\310295192\Desktop\Python\Projects\packing_list_parsing\pack.xlsx")
    sheet = wb['pack']
    x = 2
    # y = 2
    counter = 1
    while x < end_pack - 100:
        line = products[x:x + 47]

        carrot = line.split()

        # print(carrot)
        # print(carrot[0])
        # print(carrot[1])
        if 9 < counter <= 19:
            carrot[0] = "1" + carrot[0]
        elif 20 <= counter <= 29:
            carrot[0] = '2' + carrot[0]
        elif 30 <= counter:
            carrot[0] = '3' + carrot[0]
        print(carrot)
        print(carrot[0])
        print(carrot[1])
        sheet.cell(row=counter + 1, column=1).value = carrot[0]
        sheet.cell(row=counter + 1, column=2).value = carrot[1]
        sheet.cell(row=counter + 1, column=3).value = carrot[2]
        sheet.cell(row=counter + 1, column=4).value = carrot[3]
        sheet.cell(row=counter + 1, column=5).value = carrot[4]
        sheet.cell(row=counter + 1, column=6).value = carrot[5]
        sheet.cell(row=counter + 1, column=7).value = carrot[6]
        counter += 1
        x = x + 54

    if not "TOTAL" in nowastrona:
        paging = pdfreader.getPage(2)
        # wyciąga dane z pierwszej strony
        strona = paging.extractText()
        # zamienia kropki na przecinki
        nowastrona = strona.replace('.', ',')
        print(nowastrona)
        # znajduje początek listy
        start_pack = nowastrona.find('/036')

        # znajduje koniec listy
        end_pack = nowastrona.find("TOTAL")
        # tworzy zmienna tylko z danymi odnosnie paczek
        products = nowastrona[start_pack + 1:end_pack]

    x = 2
    # y = 2
    while x < end_pack - 100:
        line = products[x:x + 47]

        carrot = line.split()

        # print(carrot)
        # print(carrot[0])
        # print(carrot[1])
        if 30 <= counter < 39:
            carrot[0] = '3' + carrot[0]
        elif 40 <= counter < 49:
            carrot[0] = '4' + carrot[0]

        print(carrot)
        print(carrot[0])
        print(carrot[1])
        sheet.cell(row=counter + 1, column=1).value = carrot[0]
        sheet.cell(row=counter + 1, column=2).value = carrot[1]
        sheet.cell(row=counter + 1, column=3).value = carrot[2]
        sheet.cell(row=counter + 1, column=4).value = carrot[3]
        sheet.cell(row=counter + 1, column=5).value = carrot[4]
        sheet.cell(row=counter + 1, column=6).value = carrot[5]
        sheet.cell(row=counter + 1, column=7).value = carrot[6]
        counter += 1

        #
        x = x + 54

    file_call = str(counter - 1) + " packages"
    wb.save(r"C:\\Users\\310295192\\Desktop\\Python\\Projects\\packing_list_parsing\\" + file_call + ".xlsx")
    maximum = sheet.max_row







    #loop that is going through every line in infodis and adding nececcary data
    for x in range(1, maximum):
        # adding package quantity
        pck_quantity = browser.find_element_by_name('SL' + str(x) + 'ShipLine_Quantity')
        pck_quantity.send_keys(str(1))
        # adding type of packages
        type = browser.find_element_by_xpath("//select[@name='SL" + str(x) + "PackType_ID']/option[text()='Packages']").click()
        # adding length
        cargo = browser.find_element_by_id("cargoLineRow" + str(x))
        length = cargo.find_element_by_name('SL' + str(x) + 'ShipLine_Length')
        length.send_keys(str(sheet.cell(row=x+1, column=2).value))
        # adding width
        cargo = browser.find_element_by_id("cargoLineRow" + str(x))
        width = cargo.find_element_by_name('SL' + str(x) + 'ShipLine_Width')
        width.send_keys(str(sheet.cell(row=x+1, column=3).value))
        # adding height
        cargo = browser.find_element_by_id("cargoLineRow" + str(x))
        height = cargo.find_element_by_name('SL' + str(x) + 'ShipLine_Height')
        height.send_keys(str(sheet.cell(row=x+1, column=4).value))
        # adding volume
        cargo = browser.find_element_by_id("cargoLineRow" + str(x))
        volume = cargo.find_element_by_name('SL' + str(x) + 'ShipLine_Volume')
        volume.send_keys(str(sheet.cell(row=x+1, column=5).value))
        # adding weight
        cargo = browser.find_element_by_id("cargoLineRow" + str(x))
        weight = cargo.find_element_by_name('SL' + str(x) + 'ShipLine_Weight')
        weight.send_keys(str(sheet.cell(row=x+1, column=7).value))
        # we are adding new cargo line
        if x != maximum-1:
            buttoning = browser.find_element_by_link_text('Add cargo line')
            buttoning.click()
    wb.close()







#here i am specifing the values for canvas
canvas = tk.Canvas(root,height = wysok,width = szer)
# so pack() adds the functionality to the window
canvas.pack()
background_image =tk.PhotoImage(file = r'C:\Users\310295192\Desktop\Python\Projects\adding_packages_to_infodis\batman.png')
background_label =tk.Label(root, image = background_image)
background_label.pack(anchor = 'center')

#we can add also the color for frame
frame =tk.Frame(root,bg = 'silver', bd = 5)
#places the color in frame and indicates how much of the file it covers
frame.place(relx = 0.5, rely = 0.1, relwidth = 0.7, relheight = 0.6, anchor ='n')


# i am adding here the button specification
pack = tk.Button(frame, text = "DI & IGT Magic", font = 40, command = lambda : add_packages())
# here i am adding button to dialog window
pack.place(relx = 0.25 ,relheight = 0.3, relwidth = 0.5)

# i am adding here button to log into infodis
infodis = tk.Button(frame, text = "infodis", font = 40, command = lambda :open_infodis())
# here i am adding infodis button to frame
infodis.place (relx = 0.25,rely = 0.5, relheight = 0.3 , relwidth = 0.5)






#this lane of code stops python for further execution of code
root.mainloop()
