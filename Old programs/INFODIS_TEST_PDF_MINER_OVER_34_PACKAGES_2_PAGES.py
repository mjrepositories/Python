import PyPDF2
import openpyxl
#we are opening the file here
from tkinter import Tk
from tkinter.filedialog import askopenfilename

root = Tk()
ftypes = [('All files','*.*')]
ttl  = "Select the file"
dir1 = r'C:\Users\310295192\Desktop\Python\Projects'
root.fileName = askopenfilename(filetypes = ftypes, initialdir = dir1, title = ttl)
naming = str(root.fileName)



#otwiera plik do zapisania
file=open(naming,'rb')
#otwiera pdf
pdfreader = PyPDF2.PdfFileReader(file)
#dostaje sie do stron
print(pdfreader.getNumPages())
#otwiera pierwszą stronę
paging =pdfreader.getPage(1)
#wyciąga dane z pierwszej strony
strona = paging.extractText()
#zamienia kropki na przecinki
nowastrona=strona.replace('.',',')
print(nowastrona)
#znajduje początek listy
start_pack=nowastrona.find('/001')
print("Customer" in nowastrona)
#znalazlo na 83
#znajduje koniec listy
end_pack = nowastrona.find("Customer")
#tworzy zmienna tylko z danymi odnosnie paczek
products =nowastrona[start_pack+1:end_pack]
print(start_pack)
print(end_pack)
print(products)
wb=openpyxl.load_workbook(r"C:\Users\310295192\Desktop\Python\Projects\packing_list_parsing\pack.xlsx")
sheet= wb['pack']
x = 2
# y = 2
counter = 1
while x<end_pack-100:
    line = products[x:x+47]


    carrot = line.split()

    # print(carrot)
    # print(carrot[0])
    # print(carrot[1])
    if 9<counter <=19:
        carrot[0]="1"+carrot[0]
    elif 20<=counter<=29:
        carrot[0]='2'+carrot[0]
    elif 30<=counter:
        carrot[0]='3'+carrot[0]
    print(carrot)
    print(carrot[0])
    print(carrot[1])
    sheet.cell(row=counter+1,column=1).value = carrot[0]
    sheet.cell(row=counter+1, column=2).value = carrot[1]
    sheet.cell(row=counter+1, column=3).value = carrot[2]
    sheet.cell(row=counter+1, column=4).value = carrot[3]
    sheet.cell(row=counter+1, column=5).value = carrot[4]
    sheet.cell(row=counter+1, column=6).value = carrot[5]
    sheet.cell(row=counter+1, column=7).value = carrot[6]
    counter += 1
    x = x + 54

if not "TOTAL" in nowastrona:
    paging =pdfreader.getPage(2)
    #wyciąga dane z pierwszej strony
    strona = paging.extractText()
    #zamienia kropki na przecinki
    nowastrona=strona.replace('.',',')
    print(nowastrona)
    #znajduje początek listy
    start_pack=nowastrona.find('/036')

    #znajduje koniec listy
    end_pack = nowastrona.find("TOTAL")
    #tworzy zmienna tylko z danymi odnosnie paczek
    products =nowastrona[start_pack+1:end_pack]

x = 2
# y = 2
while x<end_pack-100:
    line = products[x:x+47]


    carrot = line.split()

    # print(carrot)
    # print(carrot[0])
    # print(carrot[1])
    if 30<=counter<39:
        carrot[0]= '3'+carrot[0]
    elif 40 <= counter < 49:
        carrot[0] = '4' + carrot[0]

    print(carrot)
    print(carrot[0])
    print(carrot[1])
    sheet.cell(row=counter + 1, column=1).value = carrot[0]
    sheet.cell(row=counter + 1, column=2).value = carrot[1]
    sheet.cell(row=counter + 1, column=3).value = carrot[2]
    sheet.cell(row=counter + 1, column=4).value = carrot[3]
    sheet.cell(row=counter + 1, column=5).value = carrot[4]
    sheet.cell(row=counter + 1, column=6).value = carrot[5]
    sheet.cell(row=counter + 1, column=7).value = carrot[6]
    counter += 1

    #
    x = x + 54

file_call = str(counter-1)+ " packages"
wb.save(r"C:\\Users\\310295192\\Desktop\\Python\\Projects\\packing_list_parsing\\"+file_call+".xlsx")
print(sheet.max_row)
wb.close()