#concept mis to download all reports from today and update them based on last day


import win32com.client
import datetime
import os
import time
import pyautogui
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import win32print, win32api,win32com.client


#getting to inbox
folder = win32com.client.Dispatch("Outlook.Application").GetNameSpace("MAPI").GetDefaultFolder(6)
#getting to infodis folder
subfolder=folder.Folders(3)
#getting to all emails
email = subfolder.Items
#checkiing the number of emails to further take this value into the loop
x=len(email)
print(x)
#sort all emails that we have in the folder
email.Sort('ReceivedTime')
first=x-4
#starts the loop through the emails
for infodis in range(5):
#checks the number of email after sorting and extract data on it
    message = email.Item(first + infodis)
    bodyofemail = message.body
    subjectofemail=message.subject
    #checks the date of the email and assing the proper date
    when = message.SentOn
    when_formated= datetime.datetime.strftime(when,'%Y-%m-%d')
    #now we will be checking each email subject and save it in proper folder
    print(subjectofemail)
    if subjectofemail == "FCL missing rates":
        attachment = message.Attachments.Item(1)
        # if the email is for FCL missing rates - it saves it in respective folder
        attachment.SaveAsFile(r'C:\Users\310295192\Desktop\Work\Rates\Sea\reports\FCL report '+when_formated+'.xlsx')
    elif subjectofemail == "Missing Rates AIR":
        attachment = message.Attachments.Item(1)
        # if the email is for AIR missing rates - it saves it in respective folder
        attachment.SaveAsFile(r'C:\Users\310295192\Desktop\Work\Rates\Air\reports\report AIR '+when_formated+'.xlsx')
    elif subjectofemail == "Missing Rates Road Europe -NL28 IT59 GB71 NL59 RO59":
        attachment = message.Attachments.Item(1)
        # if the email is for ROAD EU missing rates - it saves it in respective folder
        attachment.SaveAsFile(r'C:\Users\310295192\Desktop\Work\Rates\Road\reports\report road '+when_formated+'.xls')
    elif subjectofemail == "LCL missing rates for DB Schenker":
        attachment = message.Attachments.Item(1)
        # if the email is for LCL DB Schenker missing rates - it saves it in respective folder
        attachment.SaveAsFile(r'C:\Users\310295192\Desktop\Work\Rates\LCL\LCL reports\DB LCL '+when_formated+'.xlsx')
    elif subjectofemail == "No. of FCL shipments":
        attachment = message.Attachments.Item(1)
# if the email is for no. of shipments - it saves it in respective folder
        attachment.SaveAsFile(r'C:\Users\310295192\Desktop\Work\Rates\Daily management\No. of FCL shipments\FCL_shipments '+when_formated+'.xlsx')
    #comment

#is opening the excel file when macro is stored
os.system('start EXCEL.EXE 'r'"C:\Users\310295192\Desktop\VBA projects\processing reports infodis\updating reports.xlsm"')

time.sleep(12)
print(pyautogui.position())
#position for enabling macro Point(x=384, y=167)
pyautogui.click(384,167)
time.sleep(2)
#position for FCL upload Point(x=921, y=336)
pyautogui.click(921,336)
time.sleep(2)
#moving to point when the drag is started
pyautogui.moveTo(772,159)
time.sleep(1)
#select by dragging
pyautogui.drag(-20,-23,1)
time.sleep(3)
#confirm selection
pyautogui.typewrite(['enter'])
time.sleep(12)

#position for LCL report Point(x=916, y=523)
pyautogui.click(916,523)
#position for selection files Point(x=777, y=159)
#moving to point when the drag is started
pyautogui.moveTo(772,159)
time.sleep(1)
#select by dragging
pyautogui.drag(-20,-23,1)
time.sleep(3)
#confirm selection
pyautogui.typewrite(['enter'])
time.sleep(12)
#position for AIR report Point(x=920, y=696)
pyautogui.click(920,696)
#position for selection files Point(x=777, y=159)
#moving to point when the drag is started
pyautogui.moveTo(772,159)
time.sleep(1)
#select by dragging
pyautogui.drag(-20,-23,1)
time.sleep(3)
#confirm selection
pyautogui.typewrite(['enter'])
time.sleep(20)

#position for ROAD report Point(x=921, y=879)
pyautogui.click(920,879)
#position for selection files Point(x=777, y=159)
#moving to point when the drag is started
pyautogui.moveTo(772,159)
time.sleep(1)
#select by dragging
pyautogui.drag(-20,-23,1)
time.sleep(3)
#confirm selection
pyautogui.typewrite(['enter'])


#updating KPI report (code below)
#Open file with KPI
kpi = openpyxl.load_workbook(r"C:\Users\310295192\Desktop\Work\Rates\Daily management\daily management.xlsx")


#Go to 'data' sheet
sheet=kpi['data']
#iterate through all cells to find the first empty cell
for cell in sheet["E"]:
    if cell.value==None:
        #not first empty cell
        lastrow = cell.row
        break
#go to file with # of shipments

noship=openpyxl.load_workbook(r'C:\Users\310295192\Desktop\Work\Rates\Daily management\No. of FCL shipments\FCL_shipments '+when_formated+'.xlsx')

sheetnoship=noship['data']
#defines the number of shipments
quantity = sheetnoship.max_row-1

#open the file for missing rates
missing_rates=openpyxl.load_workbook(r'C:\Users\310295192\Desktop\Work\Rates\Sea\reports\FCL report '+when_formated+'.xlsx')


missing_rates_sh=missing_rates['data']
#defines the number of missing rates
missing_number = missing_rates_sh.max_row-1

#write the values to KPI workbook (first missing rates, second number of shipments)
sheet.cell(row=lastrow,column=5).value = missing_number
sheet.cell(row=lastrow,column=6).value = quantity

#closes all three workbooks
kpi.save(filename=r"C:\Users\310295192\Desktop\Work\Rates\Daily management\daily management.xlsx")
kpi.close()
noship.close()
missing_rates.close()

# if it is not monday it takes one day before
checking_date = datetime.datetime.today().strftime("%A")

# if it is not monday it takes one day before
if checking_date == 'Wednesday':
    fname = r'C:\Users\310295192\Desktop\Work\Rates\Daily management\daily management.xlsx'
    df = pd.read_excel(fname, sheet_name="data")
    today = datetime.datetime.today()
    two_weeks = today - datetime.timedelta(21)
    # extracting the proper dataframe from file
    fcl_rates = df[(df['Date'] <= today) & (df['Date'] >= two_weeks)]
    fcl_rates['Date'] = pd.to_datetime(fcl_rates["Date"].astype(str), format='%Y-%m-%d')
    # getting the date format and proper indexing
    fcl_rates['Date'] = fcl_rates['Date'].dt.date
    fcl_rates.index = fcl_rates['Date']
    # multiplying values by 100 to have percentages
    fcl_rates["Target"] = fcl_rates['Target'] * 100
    fcl_rates['Shipments with rates'] = fcl_rates['Shipments with rates'] * 100
    fcl_rates['Coverage'] = ((fcl_rates['All Shipments'] - fcl_rates['Shipments without rates']) / fcl_rates[
        'All Shipments']) * 100
    # assigning date table for x axis
    dates = [x.strftime('%Y-%m-%d') for x in fcl_rates['Date']]
    # assigning table with values for y axis
    values = [round(x, 3) for x in fcl_rates['Coverage']]
    # assigning values for target
    target = [95 for x in range(len(fcl_rates['Shipments with rates']))]
    coloring = []
    # checking the color for bars, if good then green and if not - red
    for x in values:
        if x > 95:
            coloring.append('#038518')
        else:
            coloring.append('#f5424b')
    # setting up figure size
    plt.figure(dpi=1024, figsize=(10, 6))
    # creating bar chart
    plt.bar(dates, values, color=coloring)
    # creating line chart
    plt.plot(dates, target, color='#031f85', linewidth=3, linestyle='--')
    # rotating ticks
    plt.xticks(dates, rotation='vertical')
    # adding title
    plt.title('FCL shipments with rates in infodis', fontsize=14)
    # setting up label for y axis
    plt.ylabel('%', fontsize=14, rotation='horizontal')
    # setting up legend and location
    plt.legend(["Target"], loc='upper center', bbox_to_anchor=(0.5, -0.25))
    # "Packing" the figure
    plt.tight_layout()
    # setting up date for file name
    x = datetime.datetime.today().strftime("%Y-%m-%d")
    # creating file name
    f_name = r'C:\Users\310295192\Desktop\Work\Rates\Daily management\DM_Graphs\DM_FCL {}.png'.format(x)
    # saving figure
    plt.savefig(f_name)
    # printing figure
    currentprinter = win32print.GetDefaultPrinter()
    win32api.ShellExecute(0, "print", f_name, currentprinter, ".", 0)
    os.startfile(r'C:\Users\310295192\Desktop\Work\Rates\Daily management\DM_Graphs')

# preparing missing rates for Panalpina

# checking the name of the today
checking_date = datetime.datetime.today().strftime("%A")
datum = ""
# if it is not monday it takes one day before
if checking_date == 'Monday':
    datum = datetime.datetime.today()
    datum = datum.strftime('%Y-%m-%d')
# creating variable for location
    location = r'C:\Users\310295192\Desktop\Work\Rates\Air\reports\report AIR {}.xlsx'.format(datum)
    # create dataframe
    df = pd.read_excel(location)
    # Indicated shipments with Panalpina as carrier
    df = df[df['Leg 2, Carrier Name'] == 'Panalpina AIR S']
    # creating only shipments that are not picked up anyhow
    report = df[df['Pending/Solved'].isna()]
    # cut columns
    columing = [x for x in range(0, 13)] + [x for x in range(22, 28)] + [38, 39,42]
    print(columing)
    panalpina_air = report.iloc[0:, columing]
    # change the format of dates
    panalpina_air['Leg 2, Pickup Planned'].dt.strftime('%Y-%m-%d')
    panalpina_air['Leg 2, Pickup Actual'].dt.strftime('%Y-%m-%d')
    panalpina_air['Leg 2, Delivery Actual'].dt.strftime('%Y-%m-%d')
    panalpina_air['Leg 3, Delivery Planned'].dt.strftime('%Y-%m-%d')
    panalpina_air['Leg 3, Delivery Actual'].dt.strftime('%Y-%m-%d')
    # create excel object in pandas with proper formatting
    filename = r'C:\Users\310295192\Desktop\Work\Rates\Air\Panalpina\Sent reports on missing lanes - Panalpina\Panalpina AIR missing rates {}.xlsx'.format(datum)
    writer = pd.ExcelWriter(filename, engine='xlsxwriter', date_format='yyyy-mm-dd', datetime_format='yyyy-mm-dd')
    # dump data to excel
    panalpina_air.to_excel(writer, sheet_name='Missing_Rates', index=False)
    # prepare worksheet for working and the formating
    workbook = writer.book
    worksheet = writer.sheets['Missing_Rates']
    blueformat = workbook.add_format({'bg_color': '#2a5fb5'})
    darkformat = workbook.add_format({'bg_color': '#012052'})
    rows = panalpina_air.shape[0]
    columns = panalpina_air.shape[1]
    # set width of the column
    worksheet.set_column('A:V', 20)
    # save file
    writer.save()
    # playing with openpyxl and formatting
    import openpyxl
    # import modules
    from openpyxl.styles import PatternFill, Font, Color

    # open worksheet
    wb = openpyxl.load_workbook(filename)
    # assign workbook and worksheet
    sheet = wb['Missing_Rates']
    # set up color for font
    ft = Font(color="FFFFFFFF")
    # table dimensions
    rows_pan = panalpina_air.shape[0] + 1
    columns_pan = panalpina_air.shape[1]
    # loop through each row
    for x in range(1, rows_pan + 1):
        for y in range(1, columns_pan + 1):
            if x % 2 == 0:
                # bright blue
                sheet.cell(row=x, column=y).fill = PatternFill(fgColor="83C8F7", fill_type="solid")
                sheet.cell(row=x, column=y).font = ft
            # dark blue
            else:
                sheet.cell(row=x, column=y).fill = PatternFill(fgColor="1192E8", fill_type="solid")
                sheet.cell(row=x, column=y).font = ft
    # save workbook
    wb.save(filename)
    wb.close()





# preparing email for what has been added and what was wrong
    outlook = win32com.client.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    #mail.To = 'DL_PCTPolandOptilo@philips.com'
    mail.To = 'GBSCTPhilips.SMB@panalpina.com; philips.prebilling@philips.com'
    mail.CC = 'peter.van.dijk@philips.com ; Herman.Houweling@panalpina.com'
    mail.Subject = 'Panalpina AIR Missing Rates ' + str(datetime.datetime.today().strftime("%Y-%m-%d"))
    mail.Body = 'Test'
    mail.HTMLBody = '<p>Dear Panalpina,</p><p>Please find attached the overview of missing rates in infodis.</p>'\
                    '<br><p>Kind Regards,</p><p>Maciej Janowski</p>' \
                    '<p>Junior Transport Specialist</p><p>Philips Polska Sp. z o.o </p>' \
                    '<p>PCT Poland</p>'
    # attaching log file
    attachment  = filename
    mail.Attachments.Add(attachment)
    # sending email
    mail.Save()