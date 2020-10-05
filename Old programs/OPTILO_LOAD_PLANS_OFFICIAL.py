
import datetime
import openpyxl
import logging
print(datetime.datetime.now())
# creating logger for actions
logger = logging.getLogger("Load Plans Log")
# setting the level for the logger
logger.setLevel('INFO')
# setting the date for logger
today = datetime.datetime.today().strftime("%Y-%m-%d")
# setting up the format for logs
formatter = logging.Formatter('%(asctime)s///%(name)s///%(levelname)s///%(message)s','%Y-%m-%d %H:%M:%S')
# creating the location of the file where logger is stored
# setting the feature regarding location of the file
location = r'C:\Users\310295192\Desktop\Python\Projects\LOAD_PLAN\loggers\Load Plans {}.log'.format(today)
file_handler = logging.FileHandler(location)
# indicating the format that we set
file_handler.setFormatter(formatter)
# combining the logger with file
logger.addHandler(file_handler)
# indicating that process is on the way
logger.info('Procedure of checking sent load plans upload has been started')

# wanted to download the report from email and open it but csv is not supported by openpyxl
# import win32com.client
# # getting to inbox
# folder = win32com.client.Dispatch("Outlook.Application").GetNameSpace("MAPI").GetDefaultFolder(6)
# # getting to infodis folder
# subfolder = folder.Folders(7)
# # getting to all emails
# email = subfolder.Items
# # checkiing the number of emails to further take this value into the loop
# x = len(email)
# print(x)
# first = x - 50 + 1
# file_for_load_plans = ""
# for loadplan in range(50):
#     # checks the number of email after sorting and extract data on it
#     message = email.Item(first + loadplan)
#     bodyofemail = message.body.upper()
#     sendermail = message.SenderEmailAddress.upper()
#     subjectofemail = message.subject.upper()
#     time_received = message.ReceivedTime
#     date_received = message.ReceivedTime
#     reporting_date = int(time_received.strftime("%d"))
#     today = int(datetime.datetime.today().strftime('%d'))
#     reporting_time = int(time_received.strftime("%H"))
#     print(reporting_date,"_______",today)
#     load_doc = "Load plans " + datetime.datetime.today().strftime("%Y-%m-%d")
#     print(load_doc)
#     # checks the MJB number in subject
#     if 'ORDER REPORT' in bodyofemail and reporting_time == 14 and reporting_date == today and len(message.Attachments) > 0:
#         #print(bodyofemail,subjectofemail,reporting)
#         print('hello')
#         for attach in message.Attachments:
#             hawb_attach = attach.FileName
#             #attach.SaveAsFile(r'C:\\Users\\310295192\\Desktop\\' + load_doc + '.csv')
#             attach.SaveAsFile(r'C:\\Users\\310295192\\Desktop\Python\\Projects\\LOAD_PLAN\\LOAD_PLANS\\'+ load_doc +'.csv')
#             file_for_load_plans = r'C:\\Users\\310295192\\Desktop\\Python\\Projects\\LOAD_PLAN\\LOAD_PLANS\\'+ load_doc +'.csv'






mjb_set = set()
logger.info("Excel file with data has been opened for verification")
# opening the excel file
wb = openpyxl.load_workbook(r'C:\Users\310295192\Desktop\Python\Projects\LOAD_PLAN\mjb_verify.xlsx')
# assigning the sheet
sheet = wb['sheet']
#finding the max row
last = sheet.max_row
# skip list allows us to indicate which MJB numbers should be skipped (because there is some issue)
skip_list =['MJB9025160','MJB9025993']
# looping through all rows
for x in range(2,last+1):
    # if cell is not empty
    if sheet.cell(row=x,column=23).value != None:
        # assign planned pick-up date
        day_extracted = datetime.datetime.date(sheet.cell(row=x,column=23).value)
        # assign current day
        current_day = datetime.datetime.date(datetime.datetime.today())
        # assign date 6 days ahead
        next_three_days = current_day + datetime.timedelta(6)
        # if date for row is withing 6 days span
        if day_extracted >= current_day and day_extracted <= next_three_days:
            planned = sheet.cell(row=x,column=9).value
            carmel =  sheet.cell(row=x,column=11).value
            if planned == "Planned" and ("Carmel" not in carmel):
                # assign number of MJB to set
                # print('yes')
                mjb_num = sheet.cell(row=x,column=49).value
                if mjb_num not in skip_list:
                    mjb_set.add(mjb_num)
        #else:
            # just say no :)
            #print('no')

# print(mjb_set)
logger.info("List with MJB number for shipments in range today + 6 days has been created")
# converting set to list
mjb_list = list(mjb_set)
logger.info(mjb_list)
# print(mjb_list)

# print(len(mjb_list))
# for x in mjb_list:
#     print(x)
wb.close()

logging.info("Checking the emails for load plans sent")
import win32com.client
import datetime
# getting to inbox
folder = win32com.client.Dispatch("Outlook.Application").GetNameSpace("MAPI").GetDefaultFolder(6)
# getting to infodis folder
subfolder = folder.Folders(7)
# getting to all emails
email = subfolder.Items
# checkiing the number of emails to further take this value into the loop
x = len(email)
# print(x)

address = set()
# sort all emails that we have in the folder
email.Sort('ReceivedTime')
first = x - 6000 + 1
# starts the loop through the emails
for load in range(6000):
    # checks the number of email after sorting and extract data on it
    message = email.Item(first + load)
    # getting the time of receiving the email
    timing = message.ReceivedTime
    # setting up the proper format email time
    day = datetime.datetime.strftime(timing,"%Y-%m-%d")
    # converting date of email to date format
    object_day = datetime.datetime.strptime(day,"%Y-%m-%d")
    # setting up the proper format for today
    today = datetime.datetime.today().strftime('%Y-%m-%d')
    # conveting date of emila to date format
    object_today = datetime.datetime.strptime(today,"%Y-%m-%d")
    # assigning body of the email
    bodyofemail = message.body
    # assigning info on sender
    sendermail = message.SenderEmailAddress.upper()
    # assigning subject of the email
    subjectofemail = message.subject.upper()
    # if day of mail is today
    #if day == today:
    # if there is load in subject of the email
    if "LOAD" in subjectofemail[:14]:
        # and if there is no "re" in subject of the email
        if "RE" not in subjectofemail[:6]:
            # and we have an attachment
            if len(message.Attachments) > 0:
                # find index for mjb
                mjb_index = subjectofemail.find('MJB')
                # extract mjb number
                mjb_no = subjectofemail[mjb_index:mjb_index+10]
                # print(subjectofemail)
                # print(mjb_no)
                address.add(mjb_no)



mjb_list_email = list(address)
logging.info("List of sent load plans for MJBs (based on emails) has been created")
logging.info(mjb_list_email)
# print(mjb_list_email)
# print(len(mjb_list_email))

mjb_present = []
mjb_absent = []

# Checking matching MJBs between two lists
for shipment in mjb_list:
    #if mjb number is in
    if shipment in mjb_list_email:
        # add to list with present MJB numbers
        mjb_present.append(shipment)
        # else - > add number to absent MJB numbers
    else:
        mjb_absent.append(shipment)


wb = openpyxl.load_workbook(r'C:\Users\310295192\Desktop\Python\Projects\LOAD_PLAN\mjb_verify.xlsx')
# assigning the sheet
sheet = wb['sheet']
#finding the max row
last = sheet.max_row
mjb_string = ""
# looping through all rows
for mjb_value in mjb_absent:
    for x in range(2,last+1):
        # if we have value for mjb in file
        if mjb_value == sheet.cell(row=x, column=49).value:
            # assigns operating user
            responsible = sheet.cell(row=x, column=3).value
            responsible = responsible.rstrip()
            # and pick up location
            pick_up = sheet.cell(row=x, column=11).value
            pick_up = pick_up.rstrip()
            # creates string on that
            mjb_string += f'<p><b>{mjb_value}-{responsible}-{pick_up}</b></p>,'
wb.close()


# splitting string by comas
mjb_string = mjb_string.split(',')
# creating set with unique values
set_mjb_string = set(mjb_string)
good_mjb = list(set_mjb_string)
if "" in good_mjb:
    good_mjb.remove("")
for x  in good_mjb:
    # print(x)
    pass
good_mjb.sort(key= lambda name: name.split("-")[1].split(" ")[0])
mjb_string_unique = ""
# parsing string to create info sent by mail
for x in good_mjb:
    mjb_string_unique += x +'\n'

print(mjb_string_unique)




# creating string for correct and wrong mjbs
correct = "Loadplans sent for below MJB numbers:\n"
wrong = "Loadplans not sent for below MJB numbers:\n"
if len(mjb_absent) != 0:
    # looping through present mjbs and adding to message
    # c = 0
    w = 0
    # for x in mjb_present:
    #     correct = correct + "<p><b>"+ x +"</b></p>"+ '\n'
    #     c += 1

    # looping through absent mjbs and adding to the message
    for x in mjb_absent:
        wrong = wrong + "<p><b>"+ x +"</b></p>"+ '\n'
        w += 1

# preparing email for what has been added and what was wrong
outlook = win32com.client.Dispatch('outlook.application')
mail = outlook.CreateItem(0)
# mail.To = 'DL_PCTPolandOptilo@philips.com'
mail.To = 'maciej.janowski@philips.com'
mail.Subject = 'Load plans sent overview - ' + str(datetime.datetime.today().strftime("%Y-%m-%d"))
mail.Body = 'Test'
mail.HTMLBody = '<p>Dear Team,</p><p>Below overview of <b>un</b>sent loadplans</p>'\
                '<p>Scope of checking - last 6000 emails and 6-days-ahead pick-up date'\
                '<p>Full log of registry in the attachment</p>'\
                +mjb_string_unique+'<p>Cheers,</p><p>Maciej Janowski</p>' \
                '<p>Junior Transport Specialist</p><p>Philips Polska Sp. z o.o </p>' \
                '<p>PCT Poland</p>' +"<p></p>"+str(w)
# attaching log file
attachment  = location
mail.Attachments.Add(attachment)
# sending email
logging.info("Mail has been generated")
mail.Send()
print(datetime.datetime.now())





































# import datetime
# import openpyxl
# import logging
# print(datetime.datetime.now())
# # creating logger for actions
# logger = logging.getLogger("Load Plans Log")
# # setting the level for the logger
# logger.setLevel('INFO')
# # setting the date for logger
# today = datetime.datetime.today().strftime("%Y-%m-%d")
# # setting up the format for logs
# formatter = logging.Formatter('%(asctime)s///%(name)s///%(levelname)s///%(message)s','%Y-%m-%d %H:%M:%S')
# # creating the location of the file where logger is stored
# # setting the feature regarding location of the file
# location = r'C:\Users\310295192\Desktop\Python\Projects\LOAD_PLAN\loggers\Load Plans {}.log'.format(today)
# file_handler = logging.FileHandler(location)
# # indicating the format that we set
# file_handler.setFormatter(formatter)
# # combining the logger with file
# logger.addHandler(file_handler)
# # indicating that process is on the way
# logger.info('Procedure of checking sent load plans upload has been started')
#
#
#
#
#
#
#
#
#
# mjb_set = set()
# logger.info("Excel file with data has been opened for verification")
# # opening the excel file
# wb = openpyxl.load_workbook(r'C:\Users\310295192\Desktop\Python\Projects\LOAD_PLAN\mjb_verify.xlsx')
# # assigning the sheet
# sheet = wb['sheet']
# #finding the max row
# last = sheet.max_row
# # looping through all rows
# for x in range(2,last+1):
#     # if cell is not empty
#     if sheet.cell(row=x,column=23).value != None:
#         # assign planned pick-up date
#         day_extracted = datetime.datetime.date(sheet.cell(row=x,column=23).value)
#         # assign current day
#         current_day = datetime.datetime.date(datetime.datetime.today())
#         # assign date 6 days ahead
#         next_three_days = current_day + datetime.timedelta(6)
#         # if date for row is withing 6 days span
#         if day_extracted >= current_day and day_extracted <= next_three_days:
#             # assign number of MJB to set
#             print('yes')
#             mjb_num = sheet.cell(row=x,column=49).value
#             mjb_set.add(mjb_num)
#         else:
#             # just say no :)
#             print('no')
#
# print(mjb_set)
# logger.info("List with MJB number for shipments in range today + 6 days has been created")
# # converting set to list
# mjb_list = list(mjb_set)
# logger.info(mjb_list)
# print(mjb_list)
#
# print(len(mjb_list))
# for x in mjb_list:
#     print(x)
# wb.close()
#
# logging.info("Checking the emails for load plans sent")
# import win32com.client
# import datetime
# # getting to inbox
# folder = win32com.client.Dispatch("Outlook.Application").GetNameSpace("MAPI").GetDefaultFolder(6)
# # getting to infodis folder
# subfolder = folder.Folders(7)
# # getting to all emails
# email = subfolder.Items
# # checkiing the number of emails to further take this value into the loop
# x = len(email)
# print(x)
#
# address = set()
# # sort all emails that we have in the folder
# email.Sort('ReceivedTime')
# first = x - 6000 + 1
# # starts the loop through the emails
# for load in range(6000):
#     # checks the number of email after sorting and extract data on it
#     message = email.Item(first + load)
#     # getting the time of receiving the email
#     timing = message.ReceivedTime
#     # setting up the proper format email time
#     day = datetime.datetime.strftime(timing,"%Y-%m-%d")
#     # converting date of email to date format
#     object_day = datetime.datetime.strptime(day,"%Y-%m-%d")
#     # setting up the proper format for today
#     today = datetime.datetime.today().strftime('%Y-%m-%d')
#     # conveting date of emila to date format
#     object_today = datetime.datetime.strptime(today,"%Y-%m-%d")
#     # assigning body of the email
#     bodyofemail = message.body
#     # assigning info on sender
#     sendermail = message.SenderEmailAddress.upper()
#     # assigning subject of the email
#     subjectofemail = message.subject.upper()
#     # if day of mail is today
#     #if day == today:
#     # if there is load in subject of the email
#     if "LOAD" in subjectofemail[:14]:
#         # and if there is no "re" in subject of the email
#         if "RE" not in subjectofemail[:6]:
#             # and we have an attachment
#             if len(message.Attachments) > 0:
#                 # find index for mjb
#                 mjb_index = subjectofemail.find('MJB')
#                 # extract mjb number
#                 mjb_no = subjectofemail[mjb_index:mjb_index+10]
#                 print(subjectofemail)
#                 print(mjb_no)
#                 address.add(mjb_no)
#
#
#
# mjb_list_email = list(address)
# logging.info("List of sent load plans for MJBs (based on emails) has been created")
# logging.info(mjb_list_email)
# print(mjb_list_email)
# print(len(mjb_list_email))
#
# mjb_present = []
# mjb_absent = []
#
# # Checking matching MJBs between two lists
# for shipment in mjb_list:
#     #if mjb number is in
#     if shipment in mjb_list_email:
#         # add to list with present MJB numbers
#         mjb_present.append(shipment)
#         # else - > add number to absent MJB numbers
#     else:
#         mjb_absent.append(shipment)
#
# # creating string for correct and wrong mjbs
# correct = "Loadplans sent for below MJB numbers:\n"
# wrong = "Loadplans not sent for below MJB numbers:\n"
# if len(mjb_present)!= 0  or len(mjb_absent) != 0:
#     # looping through present mjbs and adding to message
#     c = 0
#     w = 0
#     for x in mjb_present:
#         correct = correct + "<p><b>"+ x +"</b></p>"+ '\n'
#         c += 1
#
#     # looping through absent mjbs and adding to the message
#     for x in mjb_absent:
#         wrong = wrong + "<p><b>"+x+"</b></p>"+ '\n'
#         w += 1
#
# # preparing email for what has been added and what was wrong
# outlook = win32com.client.Dispatch('outlook.application')
# mail = outlook.CreateItem(0)
# # mail.To = 'DL_PCTPolandOptilo@philips.com'
# mail.To = 'maciej.janowski@philips.com'
# mail.Subject = 'Load plans sent overview - ' + str(datetime.datetime.today().strftime("%Y-%m-%d"))
# mail.Body = 'Test'
# mail.HTMLBody = '<p>Dear Team,</p><p>Below overview of sent loadplans</p>'\
#                 '<p>Scope of checking - last 6000 emails and 6-days-ahead pick-up date'\
#                 '<p>Full log of registry in the attachment</p>'\
#                 +correct + wrong+'<p>Cheers,</p><p>Maciej Janowski</p>' \
#                 '<p>Junior Transport Specialist</p><p>Philips Polska Sp. z o.o </p>' \
#                 '<p>PCT Poland</p>' +str(c) +"<p></p>"+str(w)
# # attaching log file
# attachment  = location
# mail.Attachments.Add(attachment)
# # sending email
# mail.Send()
# print(datetime.datetime.now())